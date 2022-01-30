from django.shortcuts import render, redirect
import json
from .decorator import restriction
import requests
import csv
from django.http import HttpResponse
from io import BytesIO
from io import StringIO
import pandas as pd
from .sorting import function
from .models import *
import pathlib
from datetime import datetime, timezone

from .forms import *
import openpyxl
import xlsxwriter
from django.contrib.auth.views import LoginView
from django.template.defaulttags import register
from django.urls import reverse_lazy
from .backup import backup
from django.http import HttpResponse

import httplib2
import googleapiclient.discovery
from googleapiclient.http import MediaFileUpload
from oauth2client.service_account import ServiceAccountCredentials
import pathlib
import pandas as pd
import psycopg2
from psycopg2 import Error
from psycopg2.extensions import ISOLATION_LEVEL_AUTOCOMMIT
import pandas as pd
from .freight import ff

def renewfreight(request):

    all = Freight.objects.all()
    for i in all:
        i.delete()

    for row in ff():
        new = Freight(forwarder='',Line='',POL='',POD='',terms='',rate='0',currencyrate='USD',period='',contract='',additional='0',currencyadd='USD',margin = '')
        new.save()

        try:
            new.forwarder = row[1]
            new.save()
        except:
            pass

        try:
            new.Line = row[1]
            new.save()
        except:
            pass

        try:
            new.POL = row[2]
            new.save()
        except:
            pass

        try:
            new.POD = row[4]
            new.save()
        except:
            pass

        try:
            new.terms = row[7]
            new.save()
        except:
            pass

        try:
            new.rate = row[8]
            new.save()
        except:
            pass

        try:
            new.period = row[9].replace('/', '.')
            new.save()
        except:
            pass

        try:
            new.contract = row[10]
            new.save()
        except:
            pass

    return redirect('OPS')

def f(request):
   user = request.user
   df_sales = pd.read_excel('/Users/a111/Desktop/Script (2).xlsx', sheet_name='Trucks')
   df_sales = pd.DataFrame(df_sales)
   count = -1
   print(df_sales)
   for i in df_sales['PO date']:
       print(i)
       count += 1
       # try:
       print(df_sales.loc[count, 'Buyer'])
       client = Empresa.objects.filter(name=df_sales.loc[count, 'Buyer'])[0]
       destination = Ports.objects.filter(port=df_sales.loc[count, 'Destination / City'])[0]
       material = Materials.objects.filter(name=df_sales.loc[count, 'Product / EN643'])[0].name

       print(df_sales.loc[count, 'Supplier'])

       proveedor = Empresa.objects.filter(name=df_sales.loc[count, 'Supplier'])[0]
       origin = Ports.objects.filter(port=df_sales.loc[count, 'Origin / City'])[0]
       material1 = Materials.objects.filter(name=df_sales.loc[count, 'Product / EN643'])[0]

       min = 0
       try:
           min = float(df_sales.loc[count, 'MT/cntr'].replace(',', '.'))
       except:
           min = float(df_sales.loc[count, 'MT/cntr'])
       cost = 0

       try:
           price = float(df_sales.loc[count, 'Sales Price (USD/Ton)'].replace(',', '.'))
       except:
           price = float(df_sales.loc[count, 'Sales Price (USD/Ton)'])


       try:
           cost = float(df_sales.loc[count, 'Purchase cost'].replace(',', '.'))
       except:
           cost = float(df_sales.loc[count, 'Purchase cost'])

       sale = SO(user=user, number=df_sales.loc[count, 'SO'], client=client, destination=destination,
                 date=df_sales.loc[count, 'SO date'], material=material, cntr=int(df_sales.loc[count, 'Cntrs']), \
                 Tons=float(str(df_sales.loc[count, 'Tons']).replace(',', '.')), min=min, cost=price,
                 currency='USD',
                 comment=str(df_sales.loc[count, 'Add. Info']) + str(
                     df_sales.loc[count, 'Order Conditions / Remarks']), \
                 cpt=str(df_sales.loc[count, 'Customers Payment Terms']), stat=True)
       sale.save()

       purchaise = PO(so=sale, number=df_sales.loc[count, 'PO'][:10], Proveedor=proveedor, Origin=origin,
                      date=df_sales.loc[count, 'PO date'], material=material1,
                      cntr=int(df_sales.loc[count, 'Cntrs']), \
                      Tons=float(str(df_sales.loc[count, 'Tons']).replace(',', '.')), price=cost, currency='EUR', \
                      spt=str(df_sales.loc[count, 'Suppliers Payment Terms']))
       purchaise.save()

       ship = Shipment(po=purchaise, number=df_sales.loc[count, 'PO'],
                       forwarder=df_sales.loc[count, 'Freight Provider'],
                       carrier=df_sales.loc[count, 'Shipping Line'], \
                       cntr=int(df_sales.loc[count, 'Cntrs']),
                       bknumber=str(df_sales.loc[count, 'Forwarder Booking Number']),
                       ETD=df_sales.loc[count, 'ETD'], \
                       ETA=df_sales.loc[count, 'ETA'], BK=True, SI=False, Magic=False, margin=0, marginEUR=0,
                       Truck=True, equip='Truck', \
                       shipinstr=df_sales.loc[count, 'VGM/Si'],link='none')
       ship.save()

       new = ShipmentRate(shipment=ship,rate=float(str(df_sales.loc[count, 'Exch rate USD/EUR']).replace(',', '.')))
       new.save()
       cntr = Containers(shipment=ship,us=request.user, number='', seal=df_sales.loc[count, 'VGM/Si'], bales=0,
                        gross=float(str(df_sales.loc[count, 'Tons']).replace(',', '.')), tara=0, vgm=0)
       cntr.save()

       cost2 = Costs(shipment=ship, name='Sale', volume=ship.po.so.cost, currency=ship.po.so.currency)
       cost2.save()
       cost3 = Costs(shipment=ship, name='Purchaise', volume=-ship.po.price, currency=ship.po.currency)
       cost3.save()

       try:
           f = -float(df_sales.loc[count, 'Freight'].replace(',','.'))
       except:
           f = -float(df_sales.loc[count, 'Freight'])

       cost = Costs(shipment=ship, name='Freight', volume=f, currency='EUR')
       cost.save()
       actualizeShip(ship.id)
       # except:
       #     pass

       return redirect('OPS')
def w(request):
   user = request.user
   df_sales = pd.read_excel('/Users/a111/Desktop/Script (1) (1).xlsx', sheet_name='Shipment')
   df_sales = pd.DataFrame(df_sales)
   count = -1
   print(df_sales)
   for i in df_sales:
       count += 1
       # try:
       print(df_sales.loc[count, 'Buyer'])
       client = Empresa.objects.filter(name=df_sales.loc[count, 'Buyer'])[0]
       destination = Ports.objects.filter(port=df_sales.loc[count, 'Destination / City'])[0]
       material = Materials.objects.filter(name=df_sales.loc[count, 'Product / EN643'])[0].name

       proveedor = Empresa.objects.filter(name=df_sales.loc[count, 'Supplier'])[0]
       origin = Ports.objects.filter(port=df_sales.loc[count, 'Origin / City'])[0]
       material1 = Materials.objects.filter(name=df_sales.loc[count, 'Product / EN643'])[0]

       min = 0
       try:
           min = float(df_sales.loc[count, 'MT/cntr'].replace(',', '.'))
       except:
           min = float(df_sales.loc[count, 'MT/cntr'])
       cost = 0

       try:
           price = float(df_sales.loc[count, 'Sales Price (USD/Ton)'].replace(',', '.'))
       except:
           price = float(df_sales.loc[count, 'Sales Price (USD/Ton)'])


       try:
           cost = float(df_sales.loc[count, 'Purchase cost'].replace(',', '.'))
       except:
           cost = float(df_sales.loc[count, 'Purchase cost'])

       sale = SO(user=user, number=df_sales.loc[count, 'SO'], client=client, destination=destination,
                 date=df_sales.loc[count, 'SO date'], material=material, cntr=int(df_sales.loc[count, 'Cntrs']), \
                 Tons=float(str(df_sales.loc[count, 'Tons']).replace(',', '.')), min=min, cost=price,
                 currency='USD',
                 comment=str(df_sales.loc[count, 'Add. Info']) + str(
                     df_sales.loc[count, 'Order Conditions / Remarks']), \
                 cpt=str(df_sales.loc[count, 'Customers Payment Terms']), stat=True)
       sale.save()

       purchaise = PO(so=sale, number=df_sales.loc[count, 'PO'][:10], Proveedor=proveedor, Origin=origin,
                      date=df_sales.loc[count, 'PO date'], material=material1,
                      cntr=int(df_sales.loc[count, 'Cntrs']), \
                      Tons=float(str(df_sales.loc[count, 'Tons']).replace(',', '.')), price=cost, currency='USD', \
                      spt=str(df_sales.loc[count, 'Suppliers Payment Terms']))
       purchaise.save()

       ship = Shipment(po=purchaise, number=df_sales.loc[count, 'PO'],
                       forwarder=df_sales.loc[count, 'Freight Provider'],
                       carrier=df_sales.loc[count, 'Shipping Line'], \
                       cntr=int(df_sales.loc[count, 'Cntrs']),
                       bknumber=str(df_sales.loc[count, 'Forwarder Booking Number']),
                       ETD=df_sales.loc[count, 'ETD'], \
                       ETA=df_sales.loc[count, 'ETA'], BK=True, SI=False, Magic=False, margin=0, marginEUR=0,
                       Truck=False, equip=str(df_sales.loc[count, 'Delivery By:']), \
                       shipinstr=df_sales.loc[count, 'VGM/Si'])
       ship.save()

       new = ShipmentRate(shipment=ship,rate=float(str(df_sales.loc[count, 'Exch rate USD/EUR']).replace(',', '.')))
       new.save()

       cost2 = Costs(shipment=ship, name='Sale', volume=ship.po.so.cost, currency=ship.po.so.currency)
       cost2.save()
       cost3 = Costs(shipment=ship, name='Purchaise', volume=-ship.po.price, currency=ship.po.currency)
       cost3.save()
       try:
           f = -float(df_sales.loc[count, 'Freight'].replace(',', '.'))
       except:
           f = -float(df_sales.loc[count, 'Freight'])
       cost = Costs(shipment=ship, name='Freight', volume=f, currency='EUR')
       cost.save()
       actualizeShip(ship.id)
       # except:
       #     pass
def z(request):
   user = request.user
   df_sales = pd.read_excel('/Users/a111/Desktop/Script (1).xlsx',sheet_name='SO')
   df_sales = pd.DataFrame(df_sales)
   count = -1
   print(df_sales)
   for i in df_sales:
       count += 1
       try:
           print(df_sales.loc[count, 'Buyer'])

           client = Empresa.objects.filter(name=df_sales.loc[count, 'Buyer'])[0]
           destination = Ports.objects.filter(port=df_sales.loc[count, 'Destination / City'])[0]
           material = Materials.objects.filter(name=df_sales.loc[count, 'Product / EN643'])[0].name

           min = 0
           try:
               min = float(df_sales.loc[count, 'MT/cntr'].replace(',', '.'))
           except:
               min = float(df_sales.loc[count, 'MT/cntr'])
           cost = 0

           try:
               cost = float(df_sales.loc[count, 'Sales Price (USD/Ton)'].replace(',', '.').replace('$', ''))
           except:
               cost = float(df_sales.loc[count, 'Sales Price (USD/Ton)'])

           sale = SO(user=user, number=df_sales.loc[count, 'SO'], client=client, destination=destination,
                     date=df_sales.loc[count, 'SO date'], material=material, cntr=int(df_sales.loc[count, 'Cntrs']),\
                     Tons=float(str(df_sales.loc[count, 'Tons']).replace(',', '.')), min=min, cost=cost, currency='USD',
                     comment=str(df_sales.loc[count, 'Add. Info']) + str(df_sales.loc[count, 'Order Conditions / Remarks']),\
                     cpt=str(df_sales.loc[count, 'Customers Payment Terms']), stat=False)
           sale.save()

       except:
           pass

def ReportMonthly(request,month):

    item = request.user
    countries = Profile.objects.filter(user=item)
    filtered = Monthly.objects.none()
    final = []

    for i in countries:
        filtered = filtered | Monthly.objects.filter(origincountry=i.country)

    for n in filtered:
        if str(n.ETD[3:]) == str(month):
            final.append(n)

    # create our spreadsheet.  I will create it in memory with a StringIO
    item = request.user

    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()

    r = SO.objects.filter(user = item)
    r = list(r)
    default = workbook.add_format({'bg_color': 'yellow'})

    # ws.write_string(0, 0, 'default format')
    #
    # fmt = copy.deepcopy(default)
    # fmt.set_bold()
    # ws.write_string(0, 1, 'bolded and yellow', fmt)

    Tonsact = models.DecimalField(max_digits=10, decimal_places=2)
    min = models.CharField(max_length=20)

    transaction = models.TextField(max_length = 50)

    Truck = models.BooleanField(default = False)

    worksheet.write(0, 0,"SO date",default)
    worksheet.write(0, 1, "PO date",default)
    worksheet.write(0, 2, "Supplier",default)
    worksheet.write(0, 3, "Client",default)
    worksheet.write(0, 4, "Origin city",default)
    worksheet.write(0, 5, "Origin country",default)
    worksheet.write(0, 6, "Dest. city",default)
    worksheet.write(0, 7, "Dest. country",default)
    worksheet.write(0, 8, "Carrier",default)
    worksheet.write(0, 9, "Forwarder",default)
    worksheet.write(0, 10, "Shipment",default)
    worksheet.write(0, 11, "BK",default)
    worksheet.write(0, 12, "Grade",default)
    worksheet.write(0, 13, "Cntr",default)
    worksheet.write(0, 14, "Tons",default)
    worksheet.write(0, 15, "Min payload",default)
    worksheet.write(0, 16, "ETD",default)
    worksheet.write(0, 17, "ETA",default)
    worksheet.write(0, 18, "Margin USD",default)
    worksheet.write(0, 19,"Margin EUR",default)
    worksheet.write(0, 20, "Log. costs USD",default)
    worksheet.write(0, 21, "Log. costs USD",default)

    worksheet.write(0, 22, "Log. costs EUR",default)
    worksheet.write(0, 23, "Log. costs EUR",default)

    worksheet.write(0, 24, "Fin. costs USD",default)
    worksheet.write(0, 25, "Fin. costs USD",default)

    worksheet.write(0, 26, "Fin. costs EUR",default)
    worksheet.write(0, 27, "Fin. costs EUR",default)

    worksheet.write(0, 28, "Exchange rate",default)
    worksheet.write(0, 29, "Equip", default)


    worksheet.write(0, 30, "Tons Actual",default)
    worksheet.write(0, 31, "Transaction",default)

    worksheet.write(0, 32, "Photos", default)

    row = 1

    for i in final:
        if i.Truck == False:

            cost = MonthlyCosts.objects.filter(monthly = i)
            fincost = FinCosts.objects.filter(monthly = i)

            logusd = '='
            logusdnames = ''

            logeur = '='
            logeurnames = ''

            finusd = '='
            finusdnames = ''

            fineur = '='
            fineurnames = ''

            for z in cost:
                if z.currency == 'USD' and z.volume > 0:
                    logusd += '+' + str(z.volume)
                    logusdnames += ',' + z.name

                if z.currency == 'USD' and z.volume < 0:
                    logusd += str(z.volume)
                    logusdnames += ',' + z.name

                if z.currency == 'EUR' and z.volume > 0:
                    logeur += '+' + str(z.volume)
                    logeurnames += ',' + z.name

                if z.currency == 'EUR' and z.volume < 0:
                    logeur += str(z.volume)
                    logeurnames += ',' + z.name

            for z in fincost:
                if z.currency == 'USD' and z.volume > 0:
                    finusd += '+' + str(z.volume)
                    finusdnames += ',' + z.name

                if z.currency == 'USD' and z.volume < 0:
                    finusd += str(z.volume)
                    finusdnames += ',' + z.name

                if z.currency == 'EUR' and z.volume > 0:
                    fineur += '+' + str(z.volume)
                    fineurnames += ',' + z.name

                if z.currency == 'EUR' and z.volume < 0:
                    fineur += str(z.volume)
                    fineurnames += ',' + z.name

            worksheet.write(row, 0, i.sodate)
            worksheet.write(row, 1, i.podate)
            worksheet.write(row, 2, i.Supplier)
            worksheet.write(row, 3, i.client)
            worksheet.write(row, 4, i.origincity)
            worksheet.write(row, 5, i.origincountry)
            worksheet.write(row, 6, i.destinationcity)
            worksheet.write(row, 7, i.destinationcountry)
            worksheet.write(row, 8, i.line)
            worksheet.write(row, 9, i.forwarder)
            worksheet.write(row, 10, i.number)
            worksheet.write(row, 11, i.bknumber)
            worksheet.write(row, 12, i.material)
            worksheet.write(row, 13, i.cntr)
            worksheet.write(row, 14, i.Tons)
            worksheet.write(row, 15, i.min)
            worksheet.write(row, 16, i.ETD)
            worksheet.write(row, 17, i.ETA)
            worksheet.write(row, 18, i.margin)
            worksheet.write(row, 19, i.marginEUR)

            worksheet.write(row, 20, logusd)
            worksheet.write(row, 21, logusdnames)

            worksheet.write(row, 22, logeur)
            worksheet.write(row, 23, logeurnames)

            worksheet.write(row, 24, finusd)
            worksheet.write(row, 25, finusdnames)

            worksheet.write(row, 26, fineur)
            worksheet.write(row, 27, fineurnames)

            worksheet.write(row, 28, MonthlyRate.objects.get(monthly = i).rate)

            worksheet.write(row, 29, i.equip)

            worksheet.write(row, 30, '0')
            worksheet.write(row, 31, '0')
            worksheet.write(row, 32, i.link)


            row += 1


    for i in final:
        if i.Truck == True:

            cost = MonthlyCosts.objects.filter(monthly=i)
            fincost = FinCosts.objects.filter(monthly=i)

            logusd = '='
            logusdnames = ''

            logeur = '='
            logeurnames = ''

            finusd = '='
            finusdnames = '='

            fineur = '='
            fineurnames = ''

            for z in cost:
                print(logusdnames)
                if z.currency == 'USD' and z.volume > 0:
                    logusd += '+' + str(z.volume)
                    logusdnames += ',' + z.name
                    # print(logusdnames)
                if z.currency == 'USD' and z.volume < 0:
                    logusd += str(z.volume)
                    logusdnames += ',' + z.name

                if z.currency == 'EUR' and z.volume > 0:
                    logeur += '+' + str(z.volume)
                    logeurnames += ',' + z.name

                if z.currency == 'EUR' and z.volume < 0:
                    logeur += str(z.volume)
                    logeurnames += ',' + z.name


            for z in fincost:
                if z.currency == 'USD' and z.volume > 0:
                    finusd += '+' + str(z.volume)
                    finusdnames += ',' + z.name

                if z.currency == 'USD' and z.volume < 0:
                    finusd += str(z.volume)
                    finusdnames += ',' + z.name

                if z.currency == 'EUR' and z.volume > 0:
                    fineur += '+' + str(z.volume)
                    fineurnames += ',' + z.name

                if z.currency == 'EUR' and z.volume < 0:
                    fineur += str(z.volume)
                    fineurnames += ',' + z.name

            worksheet.write(row, 0, i.sodate)
            worksheet.write(row, 1, i.podate)
            worksheet.write(row, 2, i.Supplier)
            worksheet.write(row, 3, i.client)
            worksheet.write(row, 4, i.origincity)
            worksheet.write(row, 5, i.origincountry)
            worksheet.write(row, 6, i.destinationcity)
            worksheet.write(row, 7, i.destinationcountry)
            worksheet.write(row, 8, i.line)
            worksheet.write(row, 9, i.forwarder)
            worksheet.write(row, 10, i.number)
            worksheet.write(row, 11, i.bknumber)
            worksheet.write(row, 12, i.material)
            worksheet.write(row, 13, i.cntr)
            worksheet.write(row, 14, i.Tons)

            worksheet.write(row, 15, i.min)
            worksheet.write(row, 16, i.ETD)
            worksheet.write(row, 17, i.ETA)
            worksheet.write(row, 18, i.margin)
            worksheet.write(row, 19, i.marginEUR)

            worksheet.write(row, 20, logusd)
            worksheet.write(row, 21, logusdnames)

            worksheet.write(row, 22, logeur)
            worksheet.write(row, 23, logeurnames)
            worksheet.write(row, 24, finusd)
            worksheet.write(row, 25, finusdnames)

            worksheet.write(row, 26, fineur)
            worksheet.write(row, 27, fineurnames)

            worksheet.write(row, 28, MonthlyRate.objects.get(monthly=i).rate)
            worksheet.write(row, 29, i.equip)
            worksheet.write(row, 30, i.Tonsact)
            worksheet.write(row, 31, i.transaction)
            worksheet.write(row, 32, i.link)

            row += 1

    workbook.close()

    # create a response
    response = HttpResponse(content_type='application/vnd.ms-excel')
    # tell the browser what the file is named
    response['Content-Disposition'] = f'attachment;filename="MonthlyReport:{month}.xlsx"'
    # put the spreadsheet data into the response
    response.write(output.getvalue())
    # return the response
    return response

def uploading(user,data,shipment_id):
    count = 0
    cntr = Containers.objects.filter(shipment_id = shipment_id)

    for i in cntr:
        nuevo = counter.objects.get(name=Shipment.objects.get(pk=shipment_id).po.Origin.country)
        nuevo.volume += -1
        nuevo.save()

        nuevo1 = counter.objects.get(name=user.username)
        nuevo1.volume += -1
        nuevo1.save()

        i.delete()

    for i in data['Number']:
        new = Containers(us = user,shipment=Shipment.objects.get(pk=shipment_id), number=str(data.loc[count, 'Number']),
                         seal=str(data.loc[count, 'Seal']),\
                         bales=int(data.loc[count, 'Bales']), gross=float(data.loc[count, 'Gross']),
                         tara=float(data.loc[count, 'Tara']),\
                         vgm=float(data.loc[count, 'Gross']) + float(data.loc[count, 'Tara']))
        new.save()

        nuevo = counter.objects.get(name=Shipment.objects.get(pk=shipment_id).po.Origin.country)
        nuevo.volume += 1
        nuevo.save()

        nuevo1 = counter.objects.get(name=user.username)
        nuevo1.volume += 1
        nuevo1.save()

        count += 1

def Upload(request,shipment_id):
    if request.method == 'POST':
        form = uploadform(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            data = pd.read_csv(file)
            user = request.user
            uploading(user,data, shipment_id)
            return redirect('ContainerViews',shipment_id)
    else:
        form = uploadform()
    return render(request,'Upload.html',{'form':form,'shipment_id':shipment_id})

def followupview(request):
    queryset = track.objects.all()
    context = {
    'queryset':queryset
    }
    return render(request,'viewfollowup.html',context)

def followupUpdate(request, follow_id):
    item = track.objects.get(pk=follow_id)
    print(item.comment)
    form = followupform(instance=item)
    if request.method == 'POST':
        form = followupform(request.POST, instance=item)
        if form.is_valid:
            form.save()
            print(item.comment)
            if (item.payment_status == True) and (item.registered == True):
                item.delete()
            return redirect('followupview')
    return render(request,'followupUpdate.html',{'form':form})

def renew(claim_id):
    new = Claims.objects.get(pk=claim_id)
    if (new.dn_currency == 'USD') and (new.cn_currency == 'USD'):
        new.profit = float(new.dn_amount) - float(new.cn_amount)
        new.save()
    elif (new.dn_currency == 'USD') and (new.cn_currency == 'EUR'):
        new.profit = float(new.dn_amount) - float(new.cn_amount)/float(new.rate)
        new.save()
    elif (new.dn_currency == 'EUR') and (new.cn_currency == 'USD'):
        new.profit = float(new.dn_amount)/float(new.rate) - float(new.cn_amount)
        new.save()
    elif (new.dn_currency == 'EUR') and (new.cn_currency == 'EUR'):
        new.profit = float(new.dn_amount)/float(new.rate) - float(new.cn_amount)/float(new.rate)
        new.save()

def ClaimCreate(request,shipment_id):
    item = Monthly.objects.get(id=shipment_id)

    if item.destinationcountry == 'Ukraine':
        form = ukrclaim()
    else:
        form = claimform()

    if request.method == 'POST':

        if item.destinationcountry == 'Ukraine':
            new = ClaimsUkr(monthly=item, humidper=0, humidton=0, impur=0, docs=False, claimsupp=0, set='')
            new.save()
            form = ukrclaim(request.POST,instance = new)

        else:
            new = Claims(Monthly=item,date='',bl='',reason='',comment='',currency='',amount=0,photos='',picCust='False',picVipa='False',settlement='False',Sent='False',cn_currency='',cn_amount=0,dn_currency='',dn_amount=0,settlement_date='',\
                         profit = 0, rate = 0, cntrs = 0, tons = 0)
            new.save()
            form = claimform(request.POST, instance = new)

        if form.is_valid():
            form.save()

            if item.destinationcountry != 'Ukraine':
                renew(new.id)
            return redirect('ClaimView', shipment_id)

    return render(request, 'formsClaim.html',{'indexform':form})

def ClaimView(request,shipment_id):
    queryset = Claims.objects.filter(Monthly_id=shipment_id)
    ukr = ClaimsUkr.objects.filter(monthly_id=shipment_id)

    mnth = Monthly.objects.get(pk=shipment_id)
    number = mnth.number

    trader = 0
    dif = 0
    general = 0
    clsup = 0
    clvipa = 0

    fin = 0
    try:
        trader = Empresa.objects.filter(name = mnth.Supplier)[0].trader
        dif = mnth.Tonsact - mnth.Tons
        general = mnth.Tons - mnth.Tonsact + ukr[0].humidton + ukr[0].impur
        clsup = mnth.po.price * general
        clvipa = mnth.po.so.cost * general
        fin = clvipa - ukr[0].claimsupp
    except:
        pass

    context = {
        'trader':trader,
        'querysets': queryset,
        'len':queryset.count(),
        'ukr':ukr,
        'len1':ukr.count,
        'number': number,
        'shipment_id':shipment_id,
        'dif':dif,
        'general':general,
        'clsup':clsup,
        'clvipa':clvipa,
        'fin':fin
    }

    return render(request, 'viewsClaim.html', context=context)

def ClaimUpdate(request,claim_id):
    try:
        item = Claims.objects.get(pk=claim_id)
        form = claimform(instance=item)
        flag = True
    except:
        item = ClaimsUkr.objects.get(pk=claim_id)
        form = ukrclaim(instance=item)
        flag = False

    if request.method == 'POST':
        if flag == True:
            form = claimform(request.POST,instance=item)
        else:
            form = ukrclaim(request.POST,instance=item)
        if form.is_valid():
            form.save()
            if flag == True:
                renew(claim_id)
                return redirect('ClaimView',item.Monthly_id)
            else:
                return redirect('ClaimView',item.monthly_id)
    return render(request, 'UpdateClaim.html', {'indexform':form})

def ClaimDelete(request,claim_id):
    try:
        item = Claims.objects.get(pk=claim_id)
        flag = True
    except:
        item = ClaimsUkr.objects.get(pk=claim_id)
        flag = False

    if request.method == 'POST':
        item.delete()
        if flag == True:
            return redirect('ClaimView',item.Monthly_id)
        else:
            return redirect('ClaimView', item.monthly_id)
    if flag == True:
        return render(request, 'ClaimConfirm.html',{'shipment_id':item.Monthly_id})
    else:
        return render(request, 'ClaimConfirm.html', {'shipment_id': item.monthly_id})

def buffer(request,shipment_id):
    item = Shipment.objects.get(id=shipment_id)
    order = PO.objects.get(id=item.po_id)
    salesorder = SO.objects.get(id=order.so_id)
    cost = salesorder.cost
    price = order.price
    costs = Costs.objects.filter(shipment_id=shipment_id)
    form = bufferform()
    city = Ports.objects.filter(port=item.po.Origin)[0].country

    containers = Containers.objects.filter(shipment=item)
    peso = 0
    for i in containers:
        peso += i.gross
    min = peso / containers.count()

    if request.method == 'POST':
        origincountry = str(Ports.objects.filter(port=order.Origin)[0].country)
        destinationcountry = str(Ports.objects.filter(port=salesorder.destination)[0].country)

        new = Monthly(po = order.number,sodate=str(salesorder.date), podate=str(order.date), Supplier=str(order.Proveedor),
                      client=str(salesorder.client),\
                      origincity=str(order.Origin), origincountry=origincountry,
                      destinationcity=str(salesorder.destination), destinationcountry=destinationcountry,\
                      number=item.number,bknumber=item.bknumber, material=salesorder.material, cntr=str(item.cntr), Tons=str(peso),
                      Tonsact=0,min=str(min), ETD=str(item.ETD),\
                      ETA=str(item.ETA), margin=0, line = item.carrier, forwarder = item.forwarder,shipinstr=item.shipinstr,equip = item.equip, link=item.link)
        new.save()

        if item.Truck == True:
            new.Truck = True
            new.save()
            container = Containers.objects.get(shipment=item)
            new.Tons = container.gross
            new.Tonsact = container.tara
            new.save()
            new.transaction = container.seal
            new.save()
            new.min = container.gross
        else:
            costs = Costs.objects.filter(shipment_id=shipment_id)
            containers = Containers.objects.filter(shipment=item)
            peso = 0
            city = Ports.objects.filter(port=item.po.Origin)[0].country
            for i in containers:
                peso += i.gross
            min = peso / containers.count()
            new.min = min
            new.Tons = peso

        rate = ShipmentRate.objects.get(shipment=item)
        newrate = MonthlyRate(monthly=new, rate=rate.rate)
        newrate.save()

        for i in costs:
            piece = MonthlyCosts(monthly=new, name=i.name, volume=i.volume, currency=i.currency)
            piece.save()

        order = PO.objects.get(id=item.po_id)
        item.delete()

        alls = Shipment.objects.filter(po=order).count() + Readiness.objects.filter(po=order).count()

        salesorder = SO.objects.get(id=order.so_id)

        dst = salesorder.destination
        if alls == 0 and salesorder.stat == True:
            salesorder.delete()

        actualize(new.id)

        new = Buffer(number=item.number,proveedor=order.Proveedor.name,Origin = order.Origin.port,carrier=item.carrier,bknumber = item.bknumber,cntr=item.cntr, ETD=item.ETD, ETA=item.ETA, comment='none')
        new.save()

        form = bufferform(request.POST)
        if form.is_valid():
            comment = form.cleaned_data['Comment']
            new.comment = comment
            new.save()
        if request.user.username[:6] == 'import':
            new.Origin = dst
            new.save()

        return redirect('Particular',city, ' ')
    return render(request, 'Buffer.html', {'form': form, 'shipment_id':shipment_id,'city':city})

def comment(request,buffer_id):
    item = Buffer.objects.get(pk=buffer_id)
    form=bufferf(instance=item)
    city = Ports.objects.filter(port=item.Origin)[0].country
    if request.method == 'POST':
        form = bufferf(request.POST,instance=item)
        if form.is_valid():
            form.save()
            return redirect('Particular', city, ' ')
    return render(request,'comment.html',{'form':form, 'buffer_id':buffer_id, 'city':city})

def DelBuffer(request,buffer_id):
    item = Buffer.objects.get(pk=buffer_id)
    city = Ports.objects.filter(port=item.Origin)[0].country
    if request.method == 'POST':
        item.delete()
        return redirect('Particular', city, ' ')
    return render(request,'Del.html',{'buffer_id':buffer_id})

@restriction
def OPS(request):
    # z(request)
    # f(request)
    # w(request)
    now = datetime.now(timezone.utc)
    m = counterupd.objects.get(index='1')

    if now.strftime("%A") == 'Monday' and m.st == False:
        all = counter.objects.all()
        for i in all:
            i.delete()

        p = Profile.objects.all()

        names = []
        for i in p:
            names.append(i.user.username)

        uniques = []
        for i in names:
            if i not in uniques:
                uniques.append(i)

        print('aa',uniques)

        cntrs = []
        for i in uniques:
            item1 = counter(name=i, volume=0)
            item1.save()

        cntrs = []
        for l in p:
            cntrs.append(l.country)

        output = []
        for x in cntrs:
            if x not in output:
                output.append(x)

        for i in output:
            itemW = counter(name=i, volume=0)
            itemW.save()

        m.st = True
        m.save()

    item = request.user
    countri = Profile.objects.filter(user=item)
    countries = []
    for i in countri:
        countries.append(str(i.country))

    filteredPO = PO.objects.none()
    X = Ports.objects.none()
    Saling = Shipment.objects.none()
    readiness = Readiness.objects.none()
    filteredBuffer = Buffer.objects.none()

    countries1 = json.dumps(countries)

    re = Readiness.objects.all()
    r = list(re)

    result = 0

    cntrs = []
    for i in Profile.objects.filter(user=request.user):
        cntrs.append(i.country)

    for n in cntrs:
        result += counter.objects.get(name = n).volume


    flag = upd.objects.get(index = '1')

    if now.hour == 12:
        flag.st == False
        flag.save()

    print('alahsdkh',now.hour)
    if flag.st == False and now.hour == 20:
        print('a')
        
        lista = {
            'Readiness':[],
            'SO':[],
            'PO':[],
            'Shipments':[],
            'Containers':[],
            'Claims':[],
            'ClaimsUkr':[],
            'Monthly':Monthly.objects.all().values(),
            'Costs': [],
            'MonthlyCosts':[],
            'FinCosts':[]
        }

        lista['Readiness'].append(['Number', 'Proveedor', 'Origin', 'Date', 'Number', 'Cntrs', 'Tons', 'Comment'])
        for i in Readiness.objects.all():
            lista['Readiness'].append([i.po.number,i.Proveedor,i.Origin,i.date,i.number,i.cntr,i.Tons,i.comment])

        lista['SO'].append(['Number', 'Client', 'Destination', 'Date', 'Material', 'Cntrs', 'Tons', 'Min', 'Sales price','Currency','Cust. payment terms','Comment','Closed','User'])
        for i in SO.objects.all():
            lista['SO'].append([i.number,i.client.name,i.destination.port,i.date,i.material,i.cntr,i.Tons,i.min,i.cost,i.currency,i.cpt,i.comment,i.stat,i.user])

        lista['PO'].append(['Number', 'Supplier', 'Origin', 'Date', 'Material', 'Cntrs', 'Tons', 'Purchaise price','Currency','Supp. payment terms'])
        for i in PO.objects.all():
            lista['PO'].append([i.number,i.Proveedor.name,i.Origin.port,i.date,i.material,i.cntr,i.Tons,i.price,i.currency,i.spt])

        lista['Shipments'].append(['Number', 'Destination', 'Client', 'Origin', 'Supplier', 'Forwarder', 'Carrier','BK', 'Material', 'Cntrs', 'ETD','ETA','Margin USD','Margin EUR','Link','comment'])
        for i in Shipment.objects.all():
            lista['Shipments'].append([i.number,i.po.so.destination,i.po.so.client,i.po.Origin,i.po.Proveedor, i.forwarder,i.carrier,i.bknumber,i.po.material,i.cntr,i.ETD,i.ETA,i.margin,i.marginEUR,i.link,i.comment])

        lista['Containers'].append(['Date','Number', 'Seal', 'Bales', 'Gross','Tara', 'VGM'])
        for i in Containers.objects.all():
            lista['Containers'].append([i.shipment.bknumber,i.number,i.seal,i.gross,i.tara, i.vgm])

        lista['Claims'].append(['Date','BK','BL Number', 'Origin Country', 'Supplier', 'Sup. Trader','Dest. Country', 'Customer','Cust. Trader',\
                                    'Grade','Reason of Claim', 'Vipa comments', 'Cntrs', 'Tons','Initial Claimed Amount', 'Currency','Photos',\
                                'Pic / Customer','Pic / VIPA', 'Sent', 'CN amount','CN currency', 'DN amount','DN currency','Set-nt Date','Profit','Rate'])

        for i in Claims.objects.all():
            lista['Claims'].append([i.date,i.bl,i.Monthly.number,i.Monthly.origincountry,i.Monthly.Supplier, Empresa.objects.filter(name=i.Monthly.Supplier)[0].trader,\
                                    i.Monthly.destinationcountry,i.Monthly.client,Empresa.objects.get(name=i.Monthly.client).trader,i.Monthly.material,i.reason,\
                                    i.comment,i.cntrs,i.tons, i.amount,i.currency,i.photos,i.picCust,i.picVipa,i.Sent,i.cn_amount,i.cn_currency,i.dn_amount,i.dn_currency,i.settlement_date,\
                                    i.profit,i.rate])

        lista['ClaimsUkr'].append(['SO/PO', 'Unloading Date', 'Trader', 'Supplier', 'Factory', 'Grade', 'Truck', 'Invoice weight','Actual weight', \
                                'Weight Difference', 'Purchase price', 'Invoice price', 'Humidity %', 'Humidity tons', 'Impurities', 'General claim tons', 'Claim to Supplier', \
                                'Claim to VIPA', 'Received Docs', 'Claimed to Supplier', 'Final claim', 'Day of settlement'])

        for i in ClaimsUkr.objects.all():
            price = -float(MonthlyCosts.objects.get(monthly=i.monthly, name='Sale').volume)
            cost = MonthlyCosts.objects.get(monthly=i.monthly, name='Purchaise').volume

            lista['ClaimsUkr'].append([i.monthly.number,i.monthly.ETA,Empresa.objects.get(name=i.monthly.Supplier).trader,i.monthly.Supplier,i.monthly.client,i.monthly.material,\
                                       i.monthly.bknumber,i.monthly.Tons,i.monthly.Tonsact,i.monthly.Tonsact - i.monthly.Tons,cost,price,i.humidper,i.humidton,i.impur,\
                                       float(i.monthly.Tons) - float(i.monthly.Tonsact) + float(i.humidton + i.impur), float(cost) * (float(i.monthly.Tons) - float(i.monthly.Tonsact) + float(i.humidton) + float(i.impur)),\
                                       float(price) * (float(i.monthly.Tons) - float(i.monthly.Tonsact) + float(i.humidton + i.impur)),float(price) * (float(i.monthly.Tons) - float(i.monthly.Tonsact) + float(i.humidton) + float(i.impur) - float(i.claimsupp)),\
                                       i.docs,i.claimsupp,i.set])

        lista['MonthlyCosts'].append(['BK','Type','Volume','Currency'])
        for i in MonthlyCosts.objects.all():
            lista['MonthlyCosts'].append([i.monthly.bknumber,i.name,i.volume,i.currency])

        lista['FinCosts'].append(['BK','Type','Volume','Currency'])
        for i in FinCosts.objects.all():
            lista['FinCosts'].append([i.monthly.bknumber,i.name,i.volume,i.currency])

        lista['Costs'].append(['BK','Type','Volume','Currency'])
        for i in FinCosts.objects.all():
            lista['Costs'].append([i.shipment.bknumber,i.name,i.volume,i.currency])

        now = datetime.now(timezone.utc)
        
        CREDENTIALS_FILE = str(pathlib.Path(
            __file__).parent.resolve()) + '/backup-334515-e0b541a9ad5d.json'  # Имя файла с закрытым ключом, вы должны подставить свое
        # Читаем ключи из файла
        credentials = ServiceAccountCredentials.from_json_keyfile_name(CREDENTIALS_FILE,
                                                                       ['https://www.googleapis.com/auth/spreadsheets',
                                                                        'https://www.googleapis.com/auth/drive'])
        httpAuth = credentials.authorize(httplib2.Http())  # Авторизуемся в системе

        driveService = googleapiclient.discovery.build('drive', 'v3',
                                                       http=httpAuth)  # Выбираем работу с Google Drive и 3 версию API
        deleter = back.objects.all()
        for i in deleter:
            delete = driveService.files().delete(fileId=i.file).execute()
            i.delete()
        for i in lista:
            print(i)
            df = pd.DataFrame(list(lista[i]))
            writer = pd.ExcelWriter(str(pathlib.Path(__file__).parent.resolve())+'/backup/'+i+'.xlsx')
            # write dataframe to excel
            df.to_excel(writer)
            # save the excel
            writer.save()
            # print(now.hour)
            # print(now.hour, now.minute, now.second)
            backup(str(i),deleter)
        flag.st = True
        flag.save()

    alpha = pd.DataFrame(list(Readiness.objects.all().values('Proveedor','Origin','date','number','cntr','Tons','comment')))
    for i in countries:
        X = X | Ports.objects.filter(country=i)

    # import section

    gamma = request.user.username[:6]
    all1 = []
    if request.user.username[:6] == 'import':

        filteredSO = SO.objects.none()
        for i in X:
            filteredSO = filteredSO | SO.objects.filter(destination=i)
            filteredBuffer = filteredBuffer | Buffer.objects.filter(Origin=i)

        for i in filteredSO:
            filteredPO = filteredPO | PO.objects.filter(so=i)

        for n in filteredPO:
            Saling = Saling | Shipment.objects.filter(po_id=n.id)

        for j in filteredPO:
            readiness = readiness | Readiness.objects.filter(po_id=j.id)

        item1 = []
        for i in filteredSO:
            item1.append(i.client.name)

        item2 = []
        for i in filteredSO:
            item2.append(i.material)

        all = item1 + item2
        print(all)

        all1 = []
        for x in all:
            if x not in all1:
                all1.append(x)
        all1 = json.dumps(all1)

        all = item1 + item2

    else:
        for i in X:
            filteredPO = filteredPO | PO.objects.filter(Origin_id=i.id)
            filteredBuffer = filteredBuffer | Buffer.objects.filter(Origin=i)

        for n in filteredPO:
            Saling = Saling | Shipment.objects.filter(po_id = n.id)

        for j in filteredPO:
            readiness = readiness | Readiness.objects.filter(po_id=j.id)

        item1 = []
        for i in filteredPO:
            item1.append(i.Proveedor.name)

        item2 = []
        for i in filteredPO:
            item2.append(i.material.name)

        all = item1 + item2

        all1 = []
        for x in all:
            if x not in all1:
                all1.append(x)
        all1 = json.dumps(all1)

    form = opsform()
    if request.method == 'POST':
        form = opsform(request.POST)
        if form.is_valid():
            number = form.cleaned_data['number']
            number1 = form.cleaned_data['number1']
        return redirect('Particular', number, number1)
    Saling = Saling.order_by('id')
    readiness = readiness.order_by('id')
    filteredBuffer = filteredBuffer.order_by('id')
    context = {
        'Shipments': Saling,
        'queryset': readiness,
        'countries': countries,
        'countries1': countries1,
        'Buffer': filteredBuffer,
        'X':X,
        'len': readiness.count(),
        'len1': Saling.count(),
        'len2': filteredBuffer.count(),
        'rates':ShipmentRate,
        'all':all1,
        'form':form,
        'result': result,
        'r':r,
        'alpha':alpha,
        'gamma':gamma
    }
    return render(request, 'OPS.html', context)

def Search(request):
    item1 = SO.objects.values_list('number', flat=True)
    item1 = list(item1)

    item2 = PO.objects.values_list('number', flat=True)
    item2 = list(item2)

    item3 = Shipment.objects.values_list('number', flat=True)
    item3 = list(item3)

    item4 = Containers.objects.values_list('number', flat=True)
    item4 = list(item4)

    item5 = Shipment.objects.values_list('bknumber', flat=True)
    item5 = list(item5)

    item6 = Monthly.objects.values_list('number', flat=True)
    item6 = list(item6)

    item7 = Monthly.objects.values_list('bknumber', flat=True)
    item7 = list(item7)

    all = item1 + item2 + item3 + item4 +item5 + item6 + item7

    all1 = json.dumps(all)
    form = searchform()
    if request.method == 'POST':
        form = searchform(request.POST)
        if form.is_valid():
            item = form.cleaned_data['number']
        return redirect('Result', item)

    context = {
        'form': form,
        'all': all1
    }
    return render(request, 'Search.html', context)

def AwailableSO(request, purchaise_id):
    usr = request.user
    order = PO.objects.get(pk=purchaise_id)

    sales = SO.objects.filter(stat=False).filter(material=order.material).filter(user=usr)
    closed = {}
    for i in sales:
        closed[i.id] = 0
        POs = PO.objects.filter(so_id = i.id)
        for p in POs:
            closed[i.id] = closed[i.id] + p.Tons
        closed[i.id] = i.Tons - closed[i.id]

    context={
        'order':order,
        'sales': sales,
        'purchaise_id': purchaise_id,
        'closed': closed
    }
    return render(request,'AwailableSO.html',context)

def Allocate(request,purchaise_id, sale_id):

    away_id = SO.objects.get(user = request.user, number = '00-0000').id
    sale = SO.objects.get(pk=sale_id)
    order = PO.objects.get(pk=purchaise_id)

    number = str(sale.number) + '-xx' + str(PO.objects.filter(so=sale).count()+1)


    registered = PO.objects.filter(so_id=sale.id)
    closed=0

    for i in registered:
        closed += i.Tons
    if request.method == 'POST':
        if order.Tons > (sale.Tons - closed):
            return redirect('Overload', purchaise_id, sale_id)
        else:

            order.number = number
            order.save()
            readiness = Readiness.objects.filter(po=order)

            for i in readiness:
                i.number = number
                i.save()

            order.so = sale
            order.save()

            return redirect('Away',away_id)
    return render(request, 'Away.html', {'id':purchaise_id, "number":number})

def Overload(request, purchaise_id, sales_id):
    order = PO.objects.get(pk=purchaise_id)
    sale = SO.objects.get(pk=sales_id)

    registered = PO.objects.filter(so_id=sales_id)
    closed=0
    for i in registered:
        closed += i.Tons
    dif = order.Tons + closed - sale.Tons
    if request.method == 'POST':
        number = str(sale.number) + '-xx' + str(PO.objects.filter(so=sale).count()+1)
        order.number = number
        order.save()
        readiness = Readiness.objects.filter(po=order)
        for i in readiness:
            i.number = number
            i.save()
        order.so = sale
        order.save()
        return redirect('SalesViews')
    return render(request,'Overload.html',{'purchaise_id':purchaise_id,'dif':dif})


def Away(request, away):
    salesorder = SO.objects.get(id=away)
    queryset = PO.objects.filter(so_id=away)

    def counter(sales_id, queryset):
        salesorder = SO.objects.get(id=sales_id)
        total = salesorder.Tons
        closed = 0
        for i in queryset:
            closed = closed + i.Tons
        return total, closed

    total, closed = counter(away, queryset)
    context = {
        'purchaises': queryset,
        'total': total,
        'closed': closed,
        'salesorder': salesorder,
        'len':queryset.count()
    }
    return render(request, 'Aways.html', context)

def Result(request, number):
    purchaises = PO.objects.none()
    shipments = Shipment.objects.none()
    readiness = Readiness.objects.none()
    result = 1
    a = 9
    total = 0

    closed = 0
    sailed = {}

    try:
        cnt = Containers.objects.get(number=number)
        result = Shipment.objects.get(pk=cnt.shipment.id)

        a = 3
    except:
        pass

    try:
        result = Shipment.objects.get(number=number)
        a = 3
    except:
        pass
    try:
        result = Shipment.objects.get(bknumber=number)
        a = 3
    except:
        pass
    try:
        result = Monthly.objects.get(number=number)
        a = 0
    except:
        pass
    try:
        result = Monthly.objects.get(bknumber=number)
        a = 0
    except:
        pass

    try:
        result = SO.objects.get(number=number)
        a = 1
        purchaises = PO.objects.filter(so_id=result.id)

        item = request.user

        POs = PO.objects.filter(so_id=result.id)
        
        for p in POs:
            a1 = Readiness.objects.filter(po=p)
            a2 = Shipment.objects.filter(po=p)
            a3 = Monthly.objects.filter(po=p.number)

            for i in a1:
                total += i.Tons

            for i in a2:
                total += i.cntr * i.po.so.min

            for i in a3:
                total += i.Tons

            a3 = Monthly.objects.filter(po=p.number)

            sailed[p.id] = 0
            for i in a3:
                sailed[p.id] += i.Tons
                
    except:
        pass

    try:
        result = PO.objects.get(number=number)
        shipments = Shipment.objects.filter(po=result)
        readiness = Readiness.objects.filter(po=result)

        a3 = Monthly.objects.filter(po = result.number)

        total = 0
        for i in a3:
            total += i.Tons

        a = 2
    except:
        pass

    context = {
        'result': result,
        'a' : a,
        'purchaises': purchaises,
        'shipments':shipments,
        'readiness': readiness,
        'len1':shipments.count(),
        'len':readiness.count(),
        'loaded':total,
        'sailed':sailed
    }
    return render(request, 'Result.html', context)
class LogIn(LoginView):
    template_name = 'login.html'
    fields = '__all__'

    def get_success_url(self):
        return reverse_lazy('OPS')


    r = Readiness.objects.all()
    r = list(r)

def report(request):
    with BytesIO() as b:
        item = request.user

        countri = Profile.objects.filter(user=item)
        countries = Ports.objects.none()
        for i in countri:
            countries = countries | Ports.objects.filter(country = i.country)

        selected = Readiness.objects.none()
        for i in countries:
            selected = selected | Readiness.objects.filter(Origin = i.port)

        df = pd.DataFrame(list(selected.values('Proveedor', 'Origin', 'date', 'number', 'cntr', 'Tons', 'comment')))
        # Use the StringIO object as the filehandle.
        writer = pd.ExcelWriter(b, engine='xlsxwriter')
        df.to_excel(writer, sheet_name='Sheet1')
        writer.save()
        # Set up the Http response.
        filename = 'Readiness.xlsx'
        response = HttpResponse(
            b.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=%s' % filename
        return response

def reportcounter(request):
    # create our spreadsheet.  I will create it in memory with a StringIO
    item = request.user

    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()

    all = counter.objects.all()


    worksheet.write(0, 0,"Name")
    worksheet.write(0, 1,"Containers loaded")

    row = 1
    for i in all:
        worksheet.write(row, 0, i.name)
        worksheet.write(row, 1, i.volume)
        row += 1
    workbook.close()

    # create a response
    response = HttpResponse(content_type='application/vnd.ms-excel')
    # tell the browser what the file is named
    response['Content-Disposition'] = 'attachment;filename="Containers.xlsx"'
    # put the spreadsheet data into the response
    response.write(output.getvalue())
    # return the response
    return response

def reportOrders(request):
    # create our spreadsheet.  I will create it in memory with a StringIO
    item = request.user

    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()

    r = SO.objects.filter(user = item)
    r = list(r)

    default = workbook.add_format({'bg_color': 'yellow'})

    # ws.write_string(0, 0, 'default format')
    #
    # fmt = copy.deepcopy(default)
    # fmt.set_bold()
    # ws.write_string(0, 1, 'bolded and yellow', fmt)

    worksheet.write(0, 0,"Number")
    worksheet.write(0, 1, "Client")
    worksheet.write(0, 2, "Destination")
    worksheet.write(0, 3, "Client's paym. terms")
    worksheet.write(0, 4, "Date")
    worksheet.write(0, 5, "Material")
    worksheet.write(0, 6, "Cntr")
    worksheet.write(0, 7, "Tons")
    worksheet.write(0, 8, "Min")
    worksheet.write(0, 9, "Cost")
    worksheet.write(0, 10, "Currency")
    worksheet.write(0, 11, "Comment")

    row = 1
    for i in r:
        worksheet.write(row, 0, i.number,default)
        worksheet.write(row, 1, i.client.name,default)
        worksheet.write(row, 2, i.destination.port,default)
        worksheet.write(row, 3, i.cpt,default)
        worksheet.write(row, 4, i.date,default)
        worksheet.write(row, 5, i.material,default)
        worksheet.write(row, 6, i.cntr,default)
        worksheet.write(row, 7, i.Tons,default)
        worksheet.write(row, 8, i.min,default)
        worksheet.write(row, 9, i.cost,default)
        worksheet.write(row, 10, i.currency,default)
        worksheet.write(row, 11, i.comment,default)

        row += 1
        for n in PO.objects.filter(so=i):
            worksheet.write(row, 0, n.number)
            worksheet.write(row, 1, n.Proveedor.name)
            worksheet.write(row, 2, n.Origin.port)
            worksheet.write(row, 3, n.spt)
            worksheet.write(row, 4, n.date)
            worksheet.write(row, 5, n.material.name)
            worksheet.write(row, 6, n.cntr)
            worksheet.write(row, 7, n.Tons)
            worksheet.write(row, 9, n.price)
            worksheet.write(row, 10, n.currency)
            row += 1
    workbook.close()

    # create a response
    response = HttpResponse(content_type='application/vnd.ms-excel')
    # tell the browser what the file is named
    response['Content-Disposition'] = 'attachment;filename="SO/PO.xlsx"'
    # put the spreadsheet data into the response
    response.write(output.getvalue())
    # return the response
    return response

def reportClaims(request):

    item = request.user
    countries = Profile.objects.filter(user=item)

    filtered = Monthly.objects.none()
    final = []

    clms = Claims.objects.none()

    for i in countries:
        filtered = filtered | Monthly.objects.filter(origincountry=i.country)

    for n in filtered:
        clms = clms | Claims.objects.filter(Monthly = n)

    # create our spreadsheet.  I will create it in memory with a StringIO
    item = request.user

    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()

    r = SO.objects.filter(user = item)
    r = list(r)
    default = workbook.add_format({'bg_color': 'yellow'})

    # ws.write_string(0, 0, 'default format')
    #
    # fmt = copy.deepcopy(default)
    # fmt.set_bold()
    # ws.write_string(0, 1, 'bolded and yellow', fmt)

    worksheet.write(0, 0,"Date")
    worksheet.write(0, 1, "BL Number")
    worksheet.write(0, 2, "Shipment")
    worksheet.write(0, 3, "Origin Country")
    worksheet.write(0, 4, "Supplier")
    worksheet.write(0, 5, "Sup. Trader")
    worksheet.write(0, 6, "Dest. Country")
    worksheet.write(0, 7, "Customer")
    worksheet.write(0, 8, "Cust. Trader")
    worksheet.write(0, 9, "Grade")
    worksheet.write(0, 10, "Reason of Claim")
    worksheet.write(0, 11, "Vipa comments")
    worksheet.write(0, 12, "Cntrs")
    worksheet.write(0, 13, "Tons")
    worksheet.write(0, 14, "Initial Claimed Amount")
    worksheet.write(0, 15, "Currency")
    worksheet.write(0, 16, "Photos")
    worksheet.write(0, 17,"Pic / Customer")
    worksheet.write(0, 18, "Pic / VIPA")
    worksheet.write(0, 19, "Sent")
    worksheet.write(0, 20, "CN amount")
    worksheet.write(0, 21, "CN currency")
    worksheet.write(0, 22, "DN amount")
    worksheet.write(0, 23, "DN currency")
    worksheet.write(0, 24, "Set-nt Date")
    worksheet.write(0, 25, "Profit")
    worksheet.write(0, 26, "Rate")
    row = 1
    for i in clms:
        worksheet.write(row, 0, i.date)
        worksheet.write(row, 1, i.bl)
        worksheet.write(row, 2, i.Monthly.number)
        worksheet.write(row, 3, i.Monthly.origincountry)
        worksheet.write(row, 4, i.Monthly.Supplier)
        worksheet.write(row, 5, Empresa.objects.filter(name=i.Monthly.Supplier)[0].trader)
        worksheet.write(row, 6, i.Monthly.destinationcountry)
        worksheet.write(row, 7, i.Monthly.client)
        worksheet.write(row, 8, Empresa.objects.filter(name=i.Monthly.client)[0].trader)
        worksheet.write(row, 9, i.Monthly.material)
        worksheet.write(row, 10, i.reason)
        worksheet.write(row, 11, i.comment)
        worksheet.write(row, 12, i.cntrs)
        worksheet.write(row, 13, i.tons)
        worksheet.write(row, 14, i.amount)
        worksheet.write(row, 15, i.currency)
        worksheet.write(row, 16, i.photos)
        worksheet.write(row, 17, i.picCust)
        worksheet.write(row, 18, i.picVipa)
        worksheet.write(row, 19, i.Sent)

        worksheet.write(row, 22, i.cn_amount)
        worksheet.write(row, 21, i.cn_currency)
        worksheet.write(row, 20, i.settlement_date)

        worksheet.write(row, 24, i.dn_amount)
        worksheet.write(row, 23, i.dn_currency)

        worksheet.write(row, 25, i.profit)
        worksheet.write(row, 26, i.rate)
        row += 1
    workbook.close()

    # create a response
    response = HttpResponse(content_type='application/vnd.ms-excel')
    # tell the browser what the file is named
    response['Content-Disposition'] = 'attachment;filename="Claims.xlsx"'
    # put the spreadsheet data into the response
    response.write(output.getvalue())
    # return the response
    return response
def reportClaimsUA(request):

    df = list(ClaimsUkr.objects.all())

    # create our spreadsheet.  I will create it in memory with a StringIO
    item = request.user

    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet()


    default = workbook.add_format({'bg_color': 'yellow'})

    # ws.write_string(0, 0, 'default format')
    #
    # fmt = copy.deepcopy(default)
    # fmt.set_bold()
    # ws.write_string(0, 1, 'bolded and yellow', fmt)

    worksheet.write(0, 0,"SO/PO")
    worksheet.write(0, 1, "Unloading Date")
    worksheet.write(0, 2, "Trader")
    worksheet.write(0, 3, "Supplier")
    worksheet.write(0, 4, "Factory")
    worksheet.write(0, 5, "Grade")
    worksheet.write(0, 6, "Truck")
    worksheet.write(0, 7, "Invoice weight")
    worksheet.write(0, 8, "Actual weight")
    worksheet.write(0, 9, "Weight Difference")
    worksheet.write(0, 10, "Purchase price")
    worksheet.write(0, 11, "Invoice price")

    worksheet.write(0, 12, "Humidity %")
    worksheet.write(0, 13, "Humidity tons")
    worksheet.write(0, 14, "Impurities")

    worksheet.write(0, 15, "General claim tons")

    worksheet.write(0, 16, "Claim to Supplier")
    worksheet.write(0, 17,"Claim to VIPA")

    worksheet.write(0, 18, "Received Docs")

    worksheet.write(0, 19, "Claimed to Supplier")

    worksheet.write(0, 20, "Final claim")
    worksheet.write(0, 21, "Day of settlement ")
    row = 1

    for i in df:
        price = -float(MonthlyCosts.objects.get(monthly = i.monthly, name = 'Sale').volume)
        cost = MonthlyCosts.objects.get(monthly = i.monthly, name = 'Purchaise').volume
        worksheet.write(row, 0, i.monthly.number)
        worksheet.write(row, 1, i.monthly.ETA)
        worksheet.write(row, 2, Empresa.objects.filter(name=i.monthly.Supplier)[0].trader)
        worksheet.write(row, 3, i.monthly.Supplier)
        worksheet.write(row, 4, i.monthly.client)
        worksheet.write(row, 5, i.monthly.material)
        worksheet.write(row, 6, i.monthly.bknumber)
        worksheet.write(row, 7, i.monthly.Tons)
        worksheet.write(row, 8, i.monthly.Tonsact)
        worksheet.write(row, 9, i.monthly.Tonsact-i.monthly.Tons)
        worksheet.write(row, 10, cost)
        worksheet.write(row, 11, price)
        worksheet.write(row, 12, i.humidper)
        worksheet.write(row, 13, i.humidton)
        worksheet.write(row, 14, i.impur)
        worksheet.write(row, 15, float(i.monthly.Tons)-float(i.monthly.Tonsact)+float(i.humidton+i.impur))
        worksheet.write(row, 16, float(cost)*(float(i.monthly.Tons)-float(i.monthly.Tonsact)+float(i.humidton)+float(i.impur)))
        worksheet.write(row, 17, float(price)*(float(i.monthly.Tons)-float(i.monthly.Tonsact)+float(i.humidton+i.impur)))
        worksheet.write(row, 18, i.docs)
        worksheet.write(row, 19, i.claimsupp)
        worksheet.write(row, 20, float(price)*(float(i.monthly.Tons)-float(i.monthly.Tonsact)+float(i.humidton)+float(i.impur)-float(i.claimsupp)))
        worksheet.write(row, 21, i.set)
        row += 1
    workbook.close()
    # create a response
    response = HttpResponse(content_type='application/vnd.ms-excel')
    # tell the browser what the file is named
    response['Content-Disposition'] = 'attachment;filename="ClaimsUA.xlsx"'
    # put the spreadsheet data into the response
    response.write(output.getvalue())
    # return the response
    return response


@register.filter
def get_item(dictionary, key):
    return dictionary.get(key)

@restriction
def Particular(request,var,var1):
    filteredPO = PO.objects.none()

    Saling = Shipment.objects.none()
    readiness = Readiness.objects.none()
    filteredBuffer = Buffer.objects.none()

    X = Ports.objects.filter(country=var)

    item = request.user
    countri = Profile.objects.filter(user=item)
    countries = []
    for i in countri:
        countries.append(str(i.country))

    final = Ports.objects.none()
    for n in countries:
        final = final | Ports.objects.filter(country=n)

    if request.user.username[:6] == 'import':
        try:
            filteredSO = SO.objects.none()
            for i in X:
                filteredSO = filteredSO | SO.objects.filter(destination=i)

            print(filteredSO)
            for i in filteredSO:
                filteredPO = filteredPO | PO.objects.filter(so=i)
                filteredBuffer = filteredBuffer | Buffer.objects.filter(Origin=i.destination.port)

            for n in filteredPO:
                Saling = Saling | Shipment.objects.filter(po_id=n.id)

            for j in filteredPO:
                readiness = readiness | Readiness.objects.filter(po_id=j.id)
        except:
            pass

        try:
            a = Empresa.objects.filter(name=var)[0]

            filteredPO = PO.objects.none()
            filteredSO = SO.objects.none()
            for i in final:
                filteredSO = filteredSO | SO.objects.filter(destination=i).filter(client=a)
                filteredBuffer = Buffer.objects.filter(Origin=i.port)


            for i in filteredSO:
                filteredPO = filteredPO | PO.objects.filter(so=i)

            Saling = Shipment.objects.none()
            readiness = Readiness.objects.none()
            filteredBuffer = Buffer.objects.none()

            for n in filteredPO:
                Saling = Saling | Shipment.objects.filter(po_id=n.id)
            for j in filteredPO:
                readiness = readiness | Readiness.objects.filter(po_id=j.id)

        except:
            pass

        try:
            a = Materials.objects.filter(name=var)[0]

            filteredPO = PO.objects.none()
            filteredSO = SO.objects.none()
            for i in final:
                filteredSO = filteredSO | SO.objects.filter(destination=i).filter(material=a.name)
                filteredBuffer = Buffer.objects.filter(Origin=i.port)

            for i in filteredSO:
                filteredPO = filteredPO | PO.objects.filter(so=i)


            Saling = Shipment.objects.none()
            readiness = Readiness.objects.none()
            filteredBuffer = Buffer.objects.none()

            for n in filteredPO:
                Saling = Saling | Shipment.objects.filter(po_id=n.id)

            for j in filteredPO:
                readiness = readiness | Readiness.objects.filter(po_id=j.id)
        except:
            pass

        try:
            a = Empresa.objects.filter(name=var1)[0]

            filteredPO = PO.objects.none()
            filteredSO = SO.objects.none()
            for i in final:
                filteredSO = filteredSO | SO.objects.filter(destination=i).filter(client=a)
                filteredBuffer = Buffer.objects.filter(Origin=i.port)

            for i in filteredSO:
                filteredPO = filteredPO | PO.objects.filter(so=i)


            Saling = Shipment.objects.none()
            readiness = Readiness.objects.none()
            filteredBuffer = Buffer.objects.none()

            for n in filteredPO:
                Saling = Saling | Shipment.objects.filter(po_id=n.id)
            for j in filteredPO:
                readiness = readiness | Readiness.objects.filter(po_id=j.id)
        except:
            pass

        try:
            a = Materials.objects.filter(name=var1)[0]

            filteredPO = PO.objects.none()
            filteredSO = SO.objects.none()
            for i in final:
                filteredSO = filteredSO | SO.objects.filter(destination=i).filter(material=a.name)
                filteredBuffer = Buffer.objects.filter(Origin=i.port)

            for i in filteredSO:
                filteredPO = filteredPO | PO.objects.filter(so=i)

            Saling = Shipment.objects.none()
            readiness = Readiness.objects.none()
            filteredBuffer = Buffer.objects.none()

            for n in filteredPO:
                Saling = Saling | Shipment.objects.filter(po_id=n.id)

            for j in filteredPO:
                readiness = readiness | Readiness.objects.filter(po_id=j.id)
        except:
            pass

        try:
            a = Empresa.objects.filter(name=var)[0]
            b = Materials.objects.filter(name=var1)[0]

            filteredPO = PO.objects.none()
            filteredSO = SO.objects.none()
            for i in final:
                filteredSO = filteredSO | SO.objects.filter(destination=i).filter(material=b.name).filter(client=a)
                filteredBuffer = Buffer.objects.filter(Origin=i.port)

            for i in filteredSO:
                filteredPO = filteredPO | PO.objects.filter(so=i)

            Saling = Shipment.objects.none()
            readiness = Readiness.objects.none()
            filteredBuffer = Buffer.objects.none()

            for n in filteredPO:
                Saling = Saling | Shipment.objects.filter(po_id=n.id)
            for j in filteredPO:
                readiness = readiness | Readiness.objects.filter(po_id=j.id)
        except:
            pass

        try:
            a = Empresa.objects.filter(name=var1)[0]
            b = Materials.objects.filter(name=var)[0]

            filteredPO = PO.objects.none()
            filteredSO = SO.objects.none()

            Saling = Shipment.objects.none()
            readiness = Readiness.objects.none()
            filteredBuffer = Buffer.objects.none()
            for i in final:
                filteredSO = filteredSO | SO.objects.filter(destination=i).filter(material=b.name).filter(client=a)
                filteredBuffer = Buffer.objects.filter(Origin=i.port)

            for i in filteredSO:
                filteredPO = filteredPO | PO.objects.filter(so=i)

            for n in filteredPO:
                Saling = Saling | Shipment.objects.filter(po_id=n.id)

            for j in filteredPO:
                readiness = readiness | Readiness.objects.filter(po_id=j.id)
        except:
            print('2')

    else:
        try:
            for i in X:
                filteredPO = filteredPO | PO.objects.filter(Origin_id=i.id)
                filteredBuffer = filteredBuffer | Buffer.objects.filter(Origin=i.port)

            for n in filteredPO:
                Saling = Saling | Shipment.objects.filter(po_id = n.id)

            for j in filteredPO:
                readiness = readiness | Readiness.objects.filter(po_id=j.id)
        except:
            pass

        try:
            a = Empresa.objects.filter(name=var)[0]
            filteredPO = PO.objects.none()
            for i in final:
                filteredPO = filteredPO | PO.objects.filter(Origin_id=i.id).filter(Proveedor=a)
                filteredBuffer = Buffer.objects.filter(Origin=i.port).filter(proveedor=a.name)

            Saling = Shipment.objects.none()
            readiness = Readiness.objects.none()
            filteredBuffer = Buffer.objects.none()

            for n in filteredPO:
                Saling = Saling | Shipment.objects.filter(po_id = n.id)
            for j in filteredPO:
                readiness = readiness | Readiness.objects.filter(po_id=j.id)
        except:
            pass

        try:
            a = Materials.objects.filter(name=var)[0]

            filteredPO = PO.objects.none()
            for i in final:
                filteredPO = filteredPO | PO.objects.filter(Origin_id=i.id).filter(material=a)
                filteredBuffer = Buffer.objects.none()
            print('dd',filteredPO)
            Saling = Shipment.objects.none()
            readiness = Readiness.objects.none()
            filteredBuffer = Buffer.objects.none()

            for n in filteredPO:
                Saling = Saling | Shipment.objects.filter(po_id = n.id)

            for j in filteredPO:
                readiness = readiness | Readiness.objects.filter(po_id=j.id)
        except:
            pass

        try:
            a = Empresa.objects.filter(name=var1)[0]

            filteredPO = PO.objects.none()
            for i in final:
                filteredPO = filteredPO | PO.objects.filter(Origin_id=i.id).filter(Proveedor=a)
                filteredBuffer = Buffer.objects.filter(Origin=i.port).filter(proveedor=a.name)

            Saling = Shipment.objects.none()
            readiness = Readiness.objects.none()
            filteredBuffer = Buffer.objects.none()

            for n in filteredPO:
                Saling = Saling | Shipment.objects.filter(po_id = n.id)
            for j in filteredPO:
                readiness = readiness | Readiness.objects.filter(po_id=j.id)
        except:
            pass

        try:
            a = Materials.objects.filter(name=var1)[0]

            filteredPO = PO.objects.none()
            for i in final:
                filteredPO = filteredPO | PO.objects.filter(Origin_id=i.id).filter(material=a)
                filteredBuffer = Buffer.objects.none()

            Saling = Shipment.objects.none()
            readiness = Readiness.objects.none()
            filteredBuffer = Buffer.objects.none()

            for n in filteredPO:
                Saling = Saling | Shipment.objects.filter(po_id = n.id)

            for j in filteredPO:
                readiness = readiness | Readiness.objects.filter(po_id=j.id)
        except:
            pass

        try:
            a = Empresa.objects.filter(name=var)[0]
            b = Materials.objects.filter(name=var1)[0]

            filteredPO = PO.objects.none()
            for i in final:
                filteredPO = filteredPO | PO.objects.filter(Origin_id=i.id).filter(Proveedor=a).filter(material=b)
                filteredBuffer = Buffer.objects.none()

            Saling = Shipment.objects.none()
            readiness = Readiness.objects.none()
            filteredBuffer = Buffer.objects.none()

            for n in filteredPO:
                Saling = Saling | Shipment.objects.filter(po_id = n.id)
            for j in filteredPO:
                readiness = readiness | Readiness.objects.filter(po_id=j.id)
        except:
            pass

        try:
            a = Empresa.objects.filter(name=var1)[0]
            b = Materials.objects.filter(name=var)[0]

            filteredPO = PO.objects.none()
            for i in final:
                filteredPO = filteredPO | PO.objects.filter(Origin_id=i.id).filter(Proveedor=a).filter(material=b)
                filteredBuffer = Buffer.objects.none()

            Saling = Shipment.objects.none()
            readiness = Readiness.objects.none()
            filteredBuffer = Buffer.objects.none()

            for n in filteredPO:
                Saling = Saling | Shipment.objects.filter(po_id = n.id)

            for j in filteredPO:
                readiness = readiness | Readiness.objects.filter(po_id=j.id)
        except:
            print('2')

    try:
        f = Readiness.objects.filter(po_id=filteredPO[0].id)
    except:
        f = 0
    Saling = Saling.order_by('id')
    readiness = readiness.order_by('id')
    filteredBuffer = filteredBuffer.order_by('id')
    context = {
        'Shipments': Saling,
        'queryset': readiness,
        'len': Saling.count(),
        'len1': readiness.count(),

        'Buffer': filteredBuffer,
        'len2': filteredBuffer.count(),
        'country':var,
        'f':f
    }
    return render(request, 'Particular.html', context)

def ParticularSO(request,var,var1):
    filteredSO = PO.objects.none()
    
    X = Ports.objects.filter(country=var)

    # try:
    for i in X:
        filteredSO = filteredSO | SO.objects.filter(destination=i).filter(stat=False).filter(user = request.user)
        print('d',filteredSO)
    # except:
    #     pass

    try:
        a = Empresa.objects.filter(name=var)[0]
        filteredSO = SO.objects.filter(client=a).filter(user = request.user)
    except:
        pass

    try:
        a = Materials.objects.filter(name=var)[0]
        filteredSO = SO.objects.filter(material=a).filter(user = request.user)
    except:
        pass

    try:
        a = Empresa.objects.filter(name=var1)[0]
        filteredSO = SO.objects.filter(client=a).filter(user = request.user)
    except:
        pass

    try:
        a = Materials.objects.filter(name=var1)[0]
        filteredSO = SO.objects.filter(material=a).filter(user = request.user)
    except:
        pass

    try:
        a = Materials.objects.filter(name=var1)[0]
        b = Empresa.objects.filter(name=var)[0]
        filteredSO = SO.objects.filter(material=a).filter(client=b).filter(user = request.user)
    except:
        pass

    try:
        a = Materials.objects.filter(name=var)[0]
        b = Empresa.objects.filter(name=var1)[0]
        filteredSO = SO.objects.filter(material=a).filter(client=b).filter(user = request.user)
    except:
        pass


    closed = {}

    for i in filteredSO:
        closed[i.id] = 0
        POs = PO.objects.filter(so_id = i.id)
        for p in POs:
            closed[i.id] = closed[i.id] + p.Tons
    context = {
        'filtered':filteredSO,
        'closed':closed
    }
    return render(request, 'ParticularSO.html', context)


def SalesCreate(request):
    usr = request.user

    proveedores = Empresa.objects.values_list('name', flat=True)
    proveedores = list(proveedores)
    proveedores = json.dumps(proveedores)

    ports = Ports.objects.values_list('port', flat=True)
    ports = list(ports)
    ports = json.dumps(ports)

    materia = Materials.objects.values_list('name', flat=True)
    materia = list(materia)
    materia = json.dumps(materia)

    form = soindex()
    if request.method == 'POST':
        form = soindex(request.POST)
        if form.is_valid():
            client = form.cleaned_data['client']
            destination = form.cleaned_data['destination']
            date = form.cleaned_data['date']
            number = form.cleaned_data['number']
            material = form.cleaned_data['material']
            cntr = form.cleaned_data['cntr']
            Tons = form.cleaned_data['Tons']
            min = form.cleaned_data['min']
            cost = form.cleaned_data['cost']
            comment = form.cleaned_data['comment']
            cpt = form.cleaned_data['cpt']
            currency = form.cleaned_data['currency']

            date = date[8] + date[9] + '.' + date[5] + date[6] + '.' + date[0] + date[1] + date[2] + date[3]
            client = Empresa.objects.filter(name=client)[0]
            destination = Ports.objects.filter(port=destination)[0]
            item = SO(user=usr,client=client,destination=destination,date=date,number=number,material=material,cntr=cntr,Tons=Tons,min=min,cost=cost,currency=currency,comment=comment,cpt=cpt)
            item.save()
            return redirect('SalesViews')
    context = {
        'form': form,
        'proveedores': proveedores,
        'ports': ports,
        'materials': materia
    }
    return render(request, 'formsSO.html', context)

def SalesUpdate(request, sales_id):

    ports = Ports.objects.values_list('port', flat=True)
    ports = list(ports)
    ports = json.dumps(ports)

    salesorder = SO.objects.get(id=sales_id)
    current = "'"+str(salesorder.destination.port)+"'"
    form = soform(instance=salesorder)
    form1 = destform()
    if request.method == 'POST':
        form1 = destform(request.POST)
        form = soform(request.POST, instance=salesorder)
        if form1.is_valid():
            dest = form1.cleaned_data['number']
            port = Ports.objects.filter(port=dest)[0]
            salesorder.destination = port
            salesorder.save()
        if form.is_valid():
            form.save()
            return redirect('SalesViews')
    context = {
        'ports':ports,
        'form': form,
        'sales_id':sales_id,
        'form1': form1,
        'current':current
    }
    return render(request, 'UpdateSO.html', context)

def SalesUpdate1(request, sales_id):

    ports1 = Ports.objects.values_list('port', flat=True)
    ports1 = list(ports1)
    ports1 = json.dumps(ports1)

    form1 = destform()
    ports = Ports.objects.all()
    salesorder = SO.objects.get(id=sales_id)
    current = "'"+str(salesorder.destination.port)+"'"
    print(current)
    form = soform(instance=salesorder)
    if request.method == 'POST':
        form1 = destform(request.POST)
        form = soform(request.POST, instance=salesorder)
        if form.is_valid():
            form.save()
            return redirect('Result',salesorder.number )
    context = {
        'ports':ports1,
        'form': form,
        'form1': form1,
        'sales_id':salesorder,
        'number':salesorder.number,
        'current':current
    }
    return render(request, 'UpdateSO1.html', context)
def SalesDelete(request, sales_id):
    item = SO.objects.get(id=sales_id)
    if request.method == 'POST':
        item.delete()
        return redirect('SalesViews')
    context = {
        'item': item
    }
    return render(request, 'DeleteSO.html', context)

def Open(request,sales_id):
    sale = SO.objects.get(pk=sales_id)
    sale.stat = False
    sale.save()
    return redirect('Result', sale.number)

def Close1(request,sales_id):
    sale = SO.objects.get(pk=sales_id)
    sale.stat = True
    sale.save()
    return redirect('Result', sale.number)

def SalesDelete1(request, sales_id):
    item = SO.objects.get(id=sales_id)
    if request.method == 'POST':
        item.delete()
        return redirect('Result',item.number)
    context = {
        'item': item
    }
    return render(request, 'DeleteSO1.html', context)

def SalesViews(request):
    item = request.user
    closed = {}
    sales = SO.objects.filter(stat=False).filter(user=item)
    total = 0
    for n in sales:
        closed[n.id] = 0
        POs = PO.objects.filter(so_id = n.id)
        total = 0
        for p in POs:
            a1 = Readiness.objects.filter(po=p)
            a2 = Shipment.objects.filter(po=p)
            a3 = Monthly.objects.filter(po=p.number)

            for i in a1:
                total += i.Tons

            for i in a2:
                total += i.cntr * i.po.so.min

            for i in a3:
                total += i.Tons

        closed[n.id] = total
        print(closed)

    moveaway = SO.objects.get(number='00-0000',user=request.user)

    POs = PO.objects.filter(so_id=moveaway.id)
    cm=0
    for p in POs:
        cm += p.Tons

    filteredSO = SO.objects.filter(user=request.user).filter(stat=False)

    item1 = []
    for i in filteredSO:
        item1.append(i.client.name)

    d = []
    for i in filteredSO:
        d.append(i.destination.country)
    d.remove('none')

    countries = []
    for x in d:
        if x not in countries:
            countries.append(x)

    item2 = []
    for i in filteredSO:
        item2.append(i.material)

    all = item1 + item2

    all1 = []
    for x in all:
        if x not in all1:
            all1.append(x)
    all1 = json.dumps(all1)
    form = opsform()
    if request.method == 'POST':
        form = opsform(request.POST)
        if form.is_valid():
            number = form.cleaned_data['number']
            number1 = form.cleaned_data['number1']
        return redirect('ParticularSO', number,number1)
    context = {
        'all': all1,
        'form': form,
        'sales': sales,
        'closed': closed,
        'countries':countries,
        'moveaway':moveaway,
        'cm':cm
    }
    return render(request, 'views.html', context)

def PurchaisesViews(request, sales_id ):
    salesorder = SO.objects.get(id=sales_id)
    queryset = PO.objects.filter(so_id=sales_id)

    item = request.user
    closed = 0
    sales = SO.objects.filter(stat=False).filter(user=item)
    total = 0
    sailed = {}
    POs = PO.objects.filter(so_id = sales_id)
    for p in POs:
        a1 = Readiness.objects.filter(po=p)
        a2 = Shipment.objects.filter(po=p)
        a3 = Monthly.objects.filter(po=p.number)

        for i in a1:
            total += i.Tons

        for i in a2:
            total += i.cntr * i.po.so.min

        for i in a3:
            total += i.Tons

        a3 = Monthly.objects.filter(po = p.number)


        sailed[p.id] = 0
        for i in a3:
            sailed[p.id] += i.Tons

        print(sailed)

    context = {
        'sailed':sailed,
        'purchaises': queryset,
        'closed': total,
        'salesorder':salesorder,
        'len':queryset.count()
    }
    return render(request, 'viewspo.html', context)

def PurchaisesViews1(request, sales_id):
    salesorder = SO.objects.get(id=sales_id)
    queryset = PO.objects.filter(so_id=sales_id)

    def counter(sales_id, queryset):
        salesorder = SO.objects.get(id=sales_id)
        total = salesorder.Tons
        closed = 0
        for i in queryset:
            closed = closed + i.Tons
        return total, closed

    total, closed = counter(sales_id, queryset)
    context = {

        'purchaises': queryset,
        'total': total,
        'closed': closed,
        'salesorder': salesorder
    }
    return render(request, 'viewspo1.html', context)

def PurchaisesCreate(request):
    usr = request.user
    proveedores = Empresa.objects.values_list('name', flat=True)
    proveedores = list(proveedores)
    proveedores = json.dumps(proveedores)

    ports = Ports.objects.values_list('port', flat=True)
    ports = list(ports)
    ports = json.dumps(ports)

    materia = Materials.objects.values_list('name', flat=True)
    materia = list(materia)
    materia = json.dumps(materia)

    indexform = index()

    SOin = ''
    Proveedor = ''
    Origin = ''

    if request.method == 'POST':
        indexform = index(request.POST)

        if indexform.is_valid():
            Proveedor = indexform.cleaned_data['Proveedor']
            Origin = indexform.cleaned_data['Origin']
            material = indexform.cleaned_data['material']
            status = indexform.cleaned_data['status']
            date = indexform.cleaned_data['date']
            cntr = indexform.cleaned_data['cntr']
            Tons = indexform.cleaned_data['Tons']
            price = indexform.cleaned_data['price']
            spt = indexform.cleaned_data['spt']
            currency = indexform.cleaned_data['currency']

            print(Proveedor)

            date = date[8]+date[9]+'.'+date[5]+date[6]+'.'+date[0]+date[1]+date[2]+date[3]

        material = Materials.objects.filter(name=material)[0]

        provider = Empresa.objects.filter(name=Proveedor)[0]
        port = Ports.objects.filter(port=Origin)[0]
        country = port.country
        profiles = Profile.objects.filter(country = country)
        if (profiles.count()) == 0 and (request.user.username[:6] != 'import'):
            return redirect('OPS')

        for i in profiles:
            if i.user.username == 'managerlatam':
                current = SO.objects.get(number='00-0000',user=i.user)
            elif i.user.username == 'managereurope':
                current = SO.objects.get(number='00-0000',user=i.user)

        if request.user.username[:6] == 'import':
            current = SO.objects.get(number='00-0000', user=request.user)

        print
        item = PO(so=current, Proveedor = provider, Origin = port,material=material,date=date, cntr=cntr, number='number', Tons=Tons, price=price,spt=spt,currency=currency)
        item.save()
        print(item.id)
        readiness = Readiness(po=item,Proveedor=item.Proveedor,Origin=item.Origin,date=date,number='Allocate',cntr=cntr,Tons=Tons, comment=status)
        readiness.save()
        print(readiness.id)
        return redirect('OPS')

    context = {
        'indexform': indexform,
        'proveedores':proveedores,
        'ports': ports,
        'materials':materia
    }
    return render(request, 'formsPO.html', context)

def CreateBuffer(request):
    form = bfr()
    if request.method == 'POST':
        form = bfr(request.POST)
        if form.is_valid():
            form.save()
        return redirect('OPS')
    return render(request,'formsBuffer.html',{'indexform':form})

def ReadinessUpdate(request, read_id):
    r=Readiness.objects.get(pk=read_id)
    order = PO.objects.get(pk=r.po_id)

    oldnumber = r.number

    form = readiness(instance=r)
    if request.method=='POST':
        form = readiness(request.POST,instance=r)
        if form.is_valid():
            form.save()

            if r.number != oldnumber:
                order.number = r.number
                order.save()

                readines = Readiness.objects.filter(po=order)
                for i in readines:
                    i.number = r.number
                    i.save()

            return redirect('OPS')
    context = {
        'form': form,
        'read_id':read_id
    }
    return render(request,'UpdateReadiness.html',context)


def ReadinessDelete(request, read_id):
    item = Readiness.objects.get(pk=read_id)
    if request.method == 'POST':
        item.delete()
        amount = Readiness.objects.filter(po_id=item.po_id).count() + Shipment.objects.filter(po_id=item.po_id).count()
        print(amount)
        if amount == 0:
            order = PO.objects.get(pk = item.po_id)
            order.delete()
        return redirect('OPS')
    return render(request, 'ConfirmationReadiness.html')

# def Move(request, purchaise_id):
#     item = PO.objects.get(pk=purchaise_id)
#     form = moveform()
#     if request.method == 'POST':
#         form = moveform(request.POST)
#         if form.is_valid():
#             newso = form.cleaned_data['newso']
#             sub = SO.objects.get(number = newso)
#             item.so = sub
#             item.save()
#             return redirect('PurchaisesUpdate', item.id)
#     return render(request, 'move.html', {'form':form, 'id':item.id})

def PurchaisesUpdate(request, purchaise_id):
    order = PO.objects.get(id=purchaise_id)
    salesorder = SO.objects.get(id = order.so_id)
    form = poform(instance=order)
    form1 = destform1()

    ports = Ports.objects.values_list('port', flat=True)
    ports = list(ports)
    ports = json.dumps(ports)

    current = "'"+str(order.Origin.port)+"'"

    if request.method == 'POST':
        form1 = destform1(request.POST)
        form = poform(request.POST, instance=order)
        if form.is_valid():
            form.save()
        if form1.is_valid():
            dest = form1.cleaned_data['number1']
            port = Ports.objects.filter(port=dest)[0]
            order.Origin = port
            order.save()

            return redirect('Result', order.number)

    context = {
        'form': form,
        'SO': order.so_id,
        'purchaise': purchaise_id,
        'number': order.number,
        'form1':form1,
        'ports': ports,
        'current': current
    }
    return render(request, 'UpdatePO.html', context)

def PurchaisesDelete(request, purchaise_id):
    item = PO.objects.get(id=purchaise_id)
    id = item.so_id
    if request.method == 'POST':
        item.delete()
        return redirect('Search')
    context = {
        'item': item,
        'id': id,
        'number':purchaise_id
    }
    return render(request, 'ConfirmationPO.html', context)
def Split(request, read_id):
    item = Readiness.objects.get(pk=read_id)
    # item = PO.objects.get(id=readiness.po_id)
    form = splitform()
    if request.method == 'POST':
        form = splitform(request.POST)
        if form.is_valid():
            current = form.cleaned_data['current']
            cntrs = form.cleaned_data['cntr']
        a = item.Tons - current
        b = item.cntr - cntrs

        new = Readiness(po=item.po,Proveedor=item.Proveedor, Origin=item.Origin, date=item.date, number=item.number, cntr=b, Tons=a, comment=item.comment)
        new.save()

        item.Tons = current
        item.cntr = cntrs
        item.save()

        return redirect('OPS')
    context = {
        'form': form
    }
    return render(request, 'Split.html', context)

def MonthlyReports(request, shipment_id):

    item = Shipment.objects.get(id=shipment_id)

    order = PO.objects.get(id=item.po_id)
    salesorder = SO.objects.get(id=order.so_id)
    cost = salesorder.cost
    price = order.price
    costs = Costs.objects.filter(shipment_id=shipment_id)
    containers = Containers.objects.filter(shipment=item)
    peso = 0
    city = Ports.objects.filter(port=item.po.Origin)[0].country
    for i in containers:
        peso += i.gross
    min = peso / containers.count()

    if request.method == 'POST':
        origincountry = str(Ports.objects.filter(port=order.Origin)[0].country)
        destinationcountry = str(Ports.objects.filter(port=salesorder.destination)[0].country)

        new = Monthly(po = order.number,sodate=str(salesorder.date), podate = str(order.date), Supplier = str(order.Proveedor), client = str(salesorder.client),\
                      origincity = str(order.Origin),origincountry=origincountry,destinationcity=str(salesorder.destination), destinationcountry=destinationcountry,\
                      number = item.number, bknumber=item.bknumber,material = salesorder.material, cntr=str(item.cntr), Tons = peso,Tonsact = 0,min = str(min),ETD = str(item.ETD),\
                      ETA=str(item.ETA), margin=0, line = item.carrier, forwarder = item.forwarder,Truck = False, transaction = '', shipinstr=item.shipinstr,equip = item.equip, link=item.link)
        new.save()



        if item.Truck == True:
            new.Truck = True
            new.save()
            container = Containers.objects.get(shipment=item)
            new.Tonsact = container.tara
            new.save()
            new.transaction = container.seal
            new.save()

        rate = ShipmentRate.objects.get(shipment=item)
        newrate=MonthlyRate(monthly=new,rate=rate.rate)
        newrate.save()

        for i in costs:
            piece = MonthlyCosts(monthly=new, name=i.name, volume=i.volume, currency=i.currency)
            piece.save()

        order = PO.objects.get(id=item.po_id)
        item.delete()

        alls = Shipment.objects.filter(po=order).count() + Readiness.objects.filter(po=order).count()

        salesorder = SO.objects.get(id=order.so_id)


        if alls == 0 and salesorder.stat == True:
            salesorder.delete()

        actualize(new.id)
        # tracking creation
        item = track(bknumber=new.bknumber, number=new.number,Supplier=new.Supplier,origincountry=new.origincountry,material=new.material,payment_status=False,\
                     registered=False,comment = '')
        item.save()



        return redirect('Particular',city, ' ')

    return render(request, 'Monthly.html',{'shipment_id':shipment_id})

def MonthlyReports1(request, shipment_id):
    item = Shipment.objects.get(id=shipment_id)

    order = PO.objects.get(id=item.po_id)
    salesorder = SO.objects.get(id=order.so_id)
    cost = salesorder.cost
    price = order.price
    costs = Costs.objects.filter(shipment_id=shipment_id)
    containers = Containers.objects.filter(shipment=item)
    peso = 0
    city = Ports.objects.filter(port=item.po.Origin)[0].country
    for i in containers:
        peso += i.gross
    min = peso / containers.count()

    if request.method == 'POST':
        origincountry = str(Ports.objects.filter(port=order.Origin).country)[0]
        destinationcountry = str(Ports.objects.filter(port=salesorder.destination).country)[0]

        new = Monthly(sodate=str(salesorder.date), podate = str(order.date), Supplier = str(order.Proveedor), client = str(salesorder.client),\
                      origincity = str(order.Origin),origincountry=origincountry,destinationcity=str(salesorder.destination), destinationcountry=destinationcountry,\
                      number = item.number, bknumber=item.bknumber,material = salesorder.material, cntr=str(order.cntr), Tons = peso,min = str(min),ETD = str(item.ETD),\
                      ETA=str(item.ETA), margin=0,line = item.carrier, forwarder = item.forwarder,shipinstr=item.shipinstr,equip = item.equip, link=item.link)
        new.save()

        rate = ShipmentRate.objects.get(shipment=item)
        newrate=MonthlyRate(monthly=new,rate=rate.rate)
        newrate.save()

        for i in costs:
            piece = MonthlyCosts(monthly=new, name=i.name, volume=-i.volume, currency=i.currency)
            piece.save()
        item.delete()

        # tracking creation
        item = track(bknumber=new.bknumber, number=new.number,Supplier=new.Supplier,origincountry=new.origincountry,material=new.material,payment_status=False,\
                     registered=False,comment = '')
        item.save()
        return redirect('Particular',city, ' ')

    return render(request, 'Monthly1.html',{'shipment_id':shipment_id})

def MonthlyView(request):
    item = request.user
    countries = Profile.objects.filter(user=item)

    filtered = Monthly.objects.none()
    final = []

    from datetime import date

    today = date.today()

    month = today.strftime("%d/%m/%Y")[3:].replace('/','.')
    print(month)

    for i in countries:
        filtered = filtered | Monthly.objects.filter(origincountry=i.country)

    for n in filtered:
        if str(n.ETD[3:]) == str(month):
            final.append(n)

    ship = []
    truck = []
    for i in final:
        if i.Truck == False:
            ship.append(i)
        else:
            truck.append(i)

    fechas = []
    for n in filtered:
        fechas.append(str(n.ETD[3:]))
        print(fechas)

    output = []
    for x in fechas:
        if x not in output:
            output.append(x)

    try:
        output.remove(month)
    except:
        pass

    context = {
        'all':ship,
        'all1': truck,
        'fechas': output,
        'month': month,
        'len': len(ship),
        'len1': len(truck)
    }
    return render(request,'MonthlyReport.html', context)

def MonthlyPart(request,month):
    filtered = Monthly.objects.none()
    final = []
    item = request.user

    countries = Profile.objects.filter(user=item)

    for i in countries:
        filtered = filtered | Monthly.objects.filter(origincountry=i.country)

    for n in filtered:
        if str(n.ETD[3:]) == str(month):
            final.append(n)

    ship = []
    truck = []
    for i in final:
        if i.Truck == False:
            ship.append(i)
        else:
            truck.append(i)

    context = {
        'all':ship,
        'all1': truck,
        'month': month,
        'len': len(ship),
        'len1': len(truck)
    }

    return render(request,'MonthlyReportPart.html', context)



def MonthlyUpdate(request, monthly_id):
    item = Monthly.objects.get(pk=monthly_id)
    form = Monthlyform(instance=item)
    if request.method == 'POST':
        form = Monthlyform(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return redirect('MonthlyView')
    context = {
        'form':form,
        'id':monthly_id
    }
    return render(request, 'MonthlyUpdate.html', context)

def MonthlyCreate(request,monthly_id):
    item = Monthly.objects.get(pk=monthly_id)

    new = Monthly(po=item.po, sodate=item.sodate, podate=item.podate, Supplier=item.Supplier,
                  client=item.client, \
                  origincity=item.origincity, origincountry=item.origincountry,
                  destinationcity=item.destinationcity, destinationcountry=item.destinationcountry, \
                  number=item.number, bknumber=item.bknumber, material=item.material, cntr=item.cntr,
                  Tons=item.Tons,Tonsact =item.Tonsact, min=item.min, ETD=item.ETD, \
                  ETA=item.ETA, margin=0, forwarder = item.forwarder,line = item.line, transaction = item.transaction,\
                  shipinstr=item.shipinstr,equip=item.equip)
    new.save()

    costs = MonthlyCosts.objects.filter(monthly=item)
    fincosts = FinCosts.objects.filter(monthly=item)

    rate = MonthlyRate.objects.get(monthly = item)
    newrate = MonthlyRate(monthly = new, rate = rate.rate)
    newrate.save()

    for i in costs:
        nuevo = MonthlyCosts(monthly = new,name = i.name, volume=i.volume,currency=i.currency)
        nuevo.save()

    for i in fincosts:
        nuevo = FinCosts(monthly = new,name = i.name, volume=i.volume,currency=i.currency)
        nuevo.save()

    actualize(new.id)
    return redirect('MonthlyView')

def MonthlyUpdate1(request, monthly_id):
    item = Monthly.objects.get(pk=monthly_id)
    form = Monthlyform(instance=item)
    if request.method == 'POST':
        form = Monthlyform(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return redirect('Result',item.number)
    context = {
        'form':form,
        'number':item.number,
        'id': item.id
    }
    return render(request, 'MonthlyUpdate1.html', context)

def MonthlyDelete(request, monthly_id):
    item = Monthly.objects.get(pk=monthly_id)
    if request.method == 'POST':
        item.delete()
        return redirect ('MonthlyView')
    return render (request, 'ConfirmationMonthly.html')

def MonthlyDelete1(request, monthly_id):
    item = Monthly.objects.get(pk=monthly_id)
    if request.method == 'POST':
        item.delete()
        return redirect ('Search')
    return render (request, 'ConfirmationMonthly1.html',{'monthly_id':monthly_id})

def ContainerViews(request, shipment_id):
    shipment = Shipment.objects.get(pk=shipment_id)
    city = Ports.objects.filter(port=shipment.po.Origin)[0].country
    if request.user.username[:6] == 'import':
        city = Ports.objects.filter(port=shipment.po.so.destination)[0].country
    queryset = Containers.objects.filter(shipment_id=shipment_id)
    context = {
        'queryset': queryset,
        'shipment_id': shipment_id,
        'city':city,
        'truck':shipment.Truck
    }
    return render(request, 'containers.html', context)

def Container(request, shipment_id):
    a = Shipment.objects.get(pk=shipment_id)
    queryset = Containers.objects.filter(shipment_id=shipment_id)
    context = {
        'queryset': queryset,
        'number': a.number,
        'shipment_id':shipment_id
    }
    return render(request, 'containers1.html', context)

def Cont(request, shipment_id):
    a = Shipment.objects.get(pk=shipment_id)
    queryset = Containers.objects.filter(shipment_id=shipment_id)
    context = {
        'queryset': queryset,
        'number': a.number,
        'shipment_id':shipment_id
    }
    return render(request, 'container.html', context)

def ContainerCreate(request, shipment_id):
    item = Shipment.objects.get(pk=shipment_id)
    form = containerindex()
    if request.method == 'POST':
        now = datetime.now(timezone.utc)
        m = counterupd.objects.get(index='1')

        form = containerindex(request.POST)
        if form.is_valid():
            number = form.cleaned_data['number']
            seal = form.cleaned_data['seal']
            bales = form.cleaned_data['bales']
            gross = form.cleaned_data['gross']
            tara = form.cleaned_data['tara']
            base = Containers(us = request.user,shipment=item,number=number,seal=seal,bales=float(bales),gross=float(gross),tara=float(tara), vgm = float(gross)+float(tara))
            base.save()

            nuevo = counter.objects.get(name = item.po.Origin.country)
            nuevo.volume += 1
            nuevo.save()

            nuevo1 = counter.objects.get(name = request.user.username)
            nuevo1.volume += 1
            nuevo1.save()

            return redirect('ContainerViews', item.id)

    return render(request,'formsContainer.html',{'form':form,'shipment_id':shipment_id})

def ContainerCreate1(request, shipment_id):
    item = Shipment.objects.get(pk=shipment_id)
    form = containerindex()

    now = datetime.now(timezone.utc)
    m = counterupd.objects.get(index='1')

    if request.method == 'POST':

        if now.strftime("%A") == 'Friday':
            m.st = False
            m.save()

        if now.strftime("%A") == 'Sunday' and m.st == False:
            all = counter.objects.all()
            for i in all:
                i.delete()
            p = Profile.objects.all()

            cntrs = []
            for i in p:
                item1 = counter(name = i.user.username,volume = 0)
                item1.save()
                cntrs.append(i.country)
            output = []

            for x in cntrs:
                if x not in output:
                    output.append(x)

            for i in output:
                itemW = counter(name = i,volume = 0)
                itemW.save()

            m.st = True
            m.save()

        form = containerindex(request.POST)

        if form.is_valid():
            number = form.cleaned_data['number']
            seal = form.cleaned_data['seal']
            bales = form.cleaned_data['bales']
            gross = form.cleaned_data['gross']
            tara = form.cleaned_data['tara']
            base = Containers(us=request.user,shipment=item,number=number,seal=seal,bales=bales,gross=gross,tara=tara, vgm = int(gross)+int(tara))
            base.save()

            nuevo = counter.objects.get(name = item.po.Origin.country)
            nuevo.volume += 1
            nuevo.save()

            nuevo1 = counter.objects.get(name = request.user.username)
            nuevo1.volume += 1
            nuevo1.save()
            return redirect('Container', item.id)

    return render(request,'formsContainer1.html',{'form':form,'shipment_id':shipment_id})

def ContainerUpdate(request, container_id):
    item = Containers.objects.get(id=container_id)


    if item.shipment.Truck == False:
        form = containerform(instance=item)
    else:
        form = containerform1(instance=item)

    if request.method == 'POST':
        if item.shipment.Truck == False:
            form = containerform(request.POST,instance=item)
        else:
            form = containerform1(request.POST,instance=item)
        if form.is_valid:
            form.save()
            item1 = Containers.objects.get(id=item.id)
            a = item1.gross
            b = item1.tara
            item1.vgm = a + b
            item1.save()
        return redirect('ContainerViews', item.shipment_id)

    context = {
        'form': form,
        'container': item.shipment_id,
        'id': container_id
    }
    return render(request, 'UpdateContainer.html', context)

def ContainerUpdate1(request, container_id):
    item = Containers.objects.get(id=container_id)
    form = containerform(instance=item)
    if request.method == 'POST':
        form = containerform(request.POST, instance=item)
        if form.is_valid:
            form.save()
            item1 = Containers.objects.get(id=item.id)
            a = item1.gross
            b = item1.tara
            item1.vgm = a + b
            item1.save()
        return redirect('Container', item.shipment_id)

    context = {
        'form': form,
        'number': item.shipment_id,
        'id': container_id
    }
    return render(request, 'UpdateContainer1.html', context)

def ContainerDelete(request, container_id):
    item = Containers.objects.get(id=container_id)
    user = item.us.username
    if request.method == 'POST':
        item.delete()

        nuevo = counter.objects.get(name=item.shipment.po.Origin.country)
        nuevo.volume += -1
        nuevo.save()

        nuevo1 = counter.objects.get(name=user)
        nuevo1.volume += -1
        nuevo1.save()

        return redirect('ContainerViews', item.shipment_id)

    context = {
        'id': item.shipment_id
    }
    return render(request, 'ConfirmationCont.html', context)

def ContainerDelete1(request, container_id):
    item = Containers.objects.get(id=container_id)
    if request.method == 'POST':
        item.delete()

        nuevo = counter.objects.get(name=item.shipment.po.Origin.country)
        nuevo.volume += -1
        nuevo.save()

        nuevo1 = counter.objects.get(name=request.user.username)
        nuevo1.volume += -1
        nuevo1.save()

        return redirect('Container', item.shipment_id)
    context = {
        'id': container_id
    }
    return render(request, 'ConfirmationCont1.html', context)

def ShipmentView(request, purchaise_id):
    order = PO.objects.get(pk=purchaise_id)
    ship = Shipment.objects.filter(po_id=purchaise_id)
    read = Readiness.objects.filter(po_id=purchaise_id)
    month = Monthly.objects.filter(po_id=purchaise_id)
    min = SO.objects.get(id=order.so_id).min
    total = order.Tons
    closed = 0
    for i in ship:
        closed = closed + i.cntr*min
    context = {
        'id': order.so_id,
        'order':order,
        'ship': ship,
        'read': read,
        'month':month,
        'closed':closed
    }

    return render(request, 'viewsship.html', context)

def ShipmentView1(request, purchaise_id):
    order = PO.objects.get(pk=purchaise_id)
    sales = SO.objects.get(pk=order.so_id)
    ship = Shipment.objects.filter(po_id=purchaise_id)
    read = Readiness.objects.filter(po_id=purchaise_id)
    min = SO.objects.get(id=order.so_id).min
    total = order.Tons
    closed = 0
    for i in ship:
        closed = closed + i.cntr*min

    context = {
        'id': order.so_id,
        'order':order,
        'ship': ship,
        'read': read,
        'closed':closed,
        'number': sales.number
    }
    return render(request, 'viewsship1.html', context)

def Close(request,sales_id):
    sale = SO.objects.get(pk=sales_id)
    if request.method == 'POST':
        sale.stat = True
        sale.save()
        return redirect('SalesViews')
    return render(request,'Close.html')

def ShipmentUpdate(request, shipment_id):
    shipment = Shipment.objects.get(id=shipment_id)
    city = Ports.objects.filter(port=shipment.po.Origin)[0].country
    form = shipmentform(instance=shipment)
    city = Ports.objects.filter(port=shipment.po.Origin)[0].country
    flag = False

    if shipment.ETD == '':
        form = shipmentform(instance=shipment)
        flag = True
    else:
        form = shipmentform1(instance=shipment)
        flag = False

    if request.method == 'POST':

        shipment = Shipment.objects.get(id=shipment_id)
        city = Ports.objects.filter(port=shipment.po.Origin)[0].country
        flag = False
        if shipment.ETD == '':
            form = shipmentform(request.POST, instance=shipment)
            flag = True
        else:
            form = shipmentform1(request.POST, instance=shipment)
            flag = False

        if request.user.username[:6]=='import':
            city = Ports.objects.filter(port = shipment.po.so.destination)[0].country
        if form.is_valid():
            form.save()
            if flag == True:
                shipment.ETD = shipment.ETD[8] + shipment.ETD[9] + '.' + shipment.ETD[5] + shipment.ETD[6] + '.' + shipment.ETD[0] + shipment.ETD[1] + shipment.ETD[2] + shipment.ETD[3]
                shipment.save()
                shipment.ETA = shipment.ETA[8] + shipment.ETA[9] + '.' + shipment.ETA[5] + shipment.ETA[6] + '.' + shipment.ETA[0] + shipment.ETA[1] + shipment.ETA[2] + shipment.ETA[3]
                shipment.save()
            return redirect('Particular',city, ' ')
    context = {
        'form': form,
        'PO': shipment.po_id,
        'Shipment': shipment_id,
        'city':city
    }
    return render(request, 'UpdateShipment.html', context)

def ShipmentUpdate1(request, shipment_id):
    shipment = Shipment.objects.get(id=shipment_id)
    form = shipmentform(instance=shipment)
    city = Ports.objects.filter(port=shipment.po.Origin)[0].country
    flag = False
    if shipment.ETD == '':
        form = shipmentform(instance=shipment)
        flag = True
    else:
        form = shipmentform1(instance=shipment)
        flag = False

    if request.method == 'POST':

        shipment = Shipment.objects.get(id=shipment_id)
        city = Ports.objects.filter(port=shipment.po.Origin)[0].country
        flag = False
        if shipment.ETD == '':
            form = shipmentform(request.POST, instance=shipment)
            flag = True
        else:
            form = shipmentform1(request.POST, instance=shipment)
            flag = False

        if form.is_valid():
            form.save()
            if flag == True:
                shipment.ETD = shipment.ETD[8] + shipment.ETD[9] + '.' + shipment.ETD[5] + shipment.ETD[6] + '.' + \
                               shipment.ETD[0] + shipment.ETD[1] + shipment.ETD[2] + shipment.ETD[3]
                shipment.save()
                shipment.ETA = shipment.ETA[8] + shipment.ETA[9] + '.' + shipment.ETA[5] + shipment.ETA[6] + '.' + \
                               shipment.ETA[0] + shipment.ETA[1] + shipment.ETA[2] + shipment.ETA[3]
                shipment.save()
            return redirect('Result', shipment.number)

    context = {
        'form': form,
        'PO': shipment.po_id,
        'Shipment': shipment_id,
        'number': shipment.number
    }
    return render(request, 'UpdateShipment1.html', context)

def ShipmentDelete(request, shipment_id):
    item = Shipment.objects.get(id=shipment_id)
    id = item.po_id
    order = PO.objects.get(id=item.po_id)
    so = SO.objects.get(id=order.so_id)
    city = Ports.objects.filter(port=item.po.Origin)[0].country
    if request.method == 'POST':
        read = Readiness(po=order,Proveedor=order.Proveedor.name,Origin=order.Origin.port,date=order.date,number=order.number, cntr = item.cntr,\
                         Tons = int(item.cntr)*so.min, comment = 'New')
        read.save()

        q = Containers.objects.filter(shipment_id = shipment_id).count()

        nuevo = counter.objects.get(name=item.po.Origin.country)
        nuevo.volume += -q
        nuevo.save()

        nuevo1 = counter.objects.get(name=request.user.username)
        nuevo1.volume += -q
        nuevo1.save()

        item.delete()
        return redirect('Particular',city, ' ')
    context = {
        'item': item,
        'id': id,
        'shipment_id': shipment_id
    }
    return render(request, 'Confirmation.html', context)

def ShipmentDelete1(request, shipment_id):
    item = Shipment.objects.get(id=shipment_id)
    id = item.po_id
    order = PO.objects.get(id=item.po_id)
    so = SO.objects.get(id=order.so_id)
    if request.method == 'POST':
        read = Readiness(po=order,Proveedor=order.Proveedor.name,Origin=order.Origin.port,date=order.date,number=order.number, cntr = item.cntr,\
                         Tons = int(item.cntr)*so.min,price=order.price, comment = 'New')
        read.save()

        q = Containers.objects.filter(shipment_id=shipment_id).count()

        nuevo = counter.objects.get(name=item.po.Origin.country)
        nuevo.volume += -q
        nuevo.save()

        nuevo1 = counter.objects.get(name=request.user.username)
        nuevo1.volume += -q
        nuevo1.save()

        item.delete()
        return redirect('Search')
    context = {
        'item': item,
        'id': id,
        'number':item.number
    }
    return render(request, 'DeleteShipment3.html', context)

def ShipmentDetail(request, shipment_id):
    item = Shipment.objects.get(id=shipment_id)
    containers = Containers.objects.filter(shipment_id=shipment_id)
    context = {
        'item': item,
        'containers': containers
    }
    return render(request, 'ShipmentDetail.html', context)

def FreightViews(request, POL, POD):
    queryset = Freight.objects.filter(POL=POL).filter(POD=POD)
    context = {
        'queryset': queryset,
        'len': len(queryset)
    }
    return render(request, 'viewfreight.html', context)
def FreightSearch(request):
    countries = Freight.objects.values_list('POL', flat=True)
    countries = list(countries)

    output = []
    for x in countries:
        if x not in output:
            output.append(x)

    countries = output
    countries = json.dumps(countries)

    countries1 = Freight.objects.values_list('POD', flat=True)
    countries1 = list(countries1)

    output = []
    for x in countries1:
        if x not in output:
            output.append(x)

    countries1 = output
    countries1 = json.dumps(countries1)

    form = freightsearchform()
    if request.method == 'POST':
        form = freightsearchform(request.POST)
        if form.is_valid():
            POL = form.cleaned_data['POL']
            POD = form.cleaned_data['POD']
        return redirect('FreightViews', POL, POD)
    context = {
        'form': form,
        'countries': countries,
        'countries1': countries1
    }
    return render(request,'SearchFreight.html', context)

def FreightCreate(request):
    ports = Ports.objects.values_list('port', flat=True)
    ports = list(ports)
    ports = json.dumps(ports)

    form = freightform()
    if request.method == 'POST':
        form = freightform(request.POST)
        if form.is_valid():
            item = form.save()
            item.period = item.period[8] + item.period[9] + '.' + item.period[5] + item.period[6] + '.' + item.period[0] + item.period[1] + item.period[2] + item.period[3]
            item.save()
        return redirect('FreightSearch')

    context = {
        'form': form,
        'countries':ports
    }
    return render(request, 'formsFreight.html', context)

def FreightUpdate(request, freight_id):
    item = Freight.objects.get(id=freight_id)
    form = freightform1(instance=item)
    if request.method == 'POST':
        form = freightform1(request.POST, instance=item)
        if form.is_valid:
            form.save()

        return redirect('FreightViews', item.POL, item.POD)

    context = {
        'form': form,
        'POL': item.POL,
        'POD': item.POD,
        'Freight':item.id
    }
    return render(request, 'UpdateFreight.html', context)
def FreightDelete(request, freight_id):
    item = Freight.objects.get(id=freight_id)

    if request.method == 'POST':
        item.delete()
        return redirect('Search')

    return render(request, 'DeleteFreight.html')



def CostCreate(request, shipment_id):
    ship = Shipment.objects.get(pk=shipment_id)
    form = indexcost()
    if request.method == 'POST':
        form = indexcost(request.POST)
        if form.is_valid():
            name = form.cleaned_data['name']
            volume = form.cleaned_data['volume']
            currency = form.cleaned_data['currency']
            item = Costs(shipment=ship,name=name,volume=volume,currency=currency)
            item.save()
            actualizeShip(shipment_id)
        return redirect ('CostViews', shipment_id)
    context = {
        'form': form,
        'shipment_id': shipment_id
    }
    return render(request, 'formsCosts.html', context)

def CostCreate1(request, shipment_id):
    ship = Shipment.objects.get(pk=shipment_id)
    form = indexcost()
    if request.method == 'POST':
        form = indexcost(request.POST)
        if form.is_valid():
            name = form.cleaned_data['name']
            volume = form.cleaned_data['volume']
            item = Costs(shipment=ship,name=name,volume=volume)
            item.save()
            actualizeShip(shipment_id)
        return redirect ('C', shipment_id)
    context = {
        'form': form,
        'shipment_id': shipment_id
    }
    return render(request, 'formsCosts1.html', context)

def actualize(monthly_id):

    cost = MonthlyCosts.objects.filter(monthly_id=monthly_id)
    fincost = FinCosts.objects.filter(monthly_id=monthly_id)

    month = Monthly.objects.get(pk=monthly_id)

    rate = MonthlyRate.objects.get(monthly=month)
    a = 0
    for i in cost:
        if i.name == 'Sale' and i.currency == 'USD':
            a += float(i.volume)
        elif i.name == 'Sale' and i.currency == 'EUR':
            a += float(i.volume)/float(rate.rate)

        elif i.name == 'Purchaise' and i.currency == 'USD':
            a += float(i.volume)

        elif i.name == 'Purchaise' and i.currency == 'EUR':
            a += float(i.volume)/float(rate.rate)

        elif i.currency == 'USD':
            a += float(i.volume) / float(month.min)

        elif i.currency == 'EUR':
            a += ((float(i.volume) / float(month.min)))/float(rate.rate)

    for i in fincost:

        if i.currency == 'USD':
            a += float(i.volume) / float(month.min)

        if i.currency == 'EUR':
            a += ((float(i.volume) / float(month.min))) / float(rate.rate)

    month.margin = float('{:.2f}'.format(a))
    month.save()

    result = a*float(rate.rate)

    month.marginEUR = float('{:.2f}'.format(result))
    month.save()

def CostMonthlyCreate(request, monthly_id):
    month = Monthly.objects.get(pk=monthly_id)
    form = indexcost()
    a = 0
    if request.method == 'POST':
        form = indexcost(request.POST)
        if form.is_valid():
            name = form.cleaned_data['name']
            volume = form.cleaned_data['volume']
            currency = form.cleaned_data['currency']
            item = MonthlyCosts(monthly=month,name=name,volume=volume, currency=currency)
            item.save()
            actualize(item.monthly_id)
        return redirect ('CostMonthlyView', monthly_id)
    context = {
        'form': form,
    }
    return render(request, 'formsCostMonthly.html', context)
def CostMonthlyDelete(request, cost_id):
    item = MonthlyCosts.objects.get(id = cost_id)
    month = Monthly.objects.get(pk=item.monthly_id)
    if request.method == 'POST':
        item.delete()
        actualize(item.monthly_id)
        return redirect ('CostMonthlyView', month.id)
    return render(request, 'ConfirmationMonthly.html',{'id': month.id})
def CostMonthlyView(request, monthly_id):
    costs = MonthlyCosts.objects.filter(monthly_id = monthly_id)
    return render(request, 'viewscostsMonthly.html', {'costs': costs, 'monthly_id': monthly_id})

def CostMonthlyCreate1(request, monthly_id):
    month = Monthly.objects.get(pk=monthly_id)
    form = indexcost()
    a = 0
    if request.method == 'POST':
        form = indexcost(request.POST)
        if form.is_valid():
            name = form.cleaned_data['name']
            volume = form.cleaned_data['volume']
            currency = form.cleaned_data['currency']
            item = MonthlyCosts(monthly=month,name=name,volume=volume, currency=currency)
            item.save()
            actualize(item.monthly_id)
        return redirect ('CostMonthlyView1', monthly_id)
    context = {
        'form': form,
        'id': monthly_id
    }
    return render(request, 'formsCostMonthly1.html', context)

def CostMonthlyDelete1(request, cost_id):
    item = MonthlyCosts.objects.get(id = cost_id)
    month = Monthly.objects.get(pk=item.monthly_id)
    if request.method == 'POST':
        item.delete()
        actualize(item.monthly_id)
        return redirect ('CostMonthlyView1', month.id)
    return render(request, 'ConfirmationMonthlyCost1.html',{'id': month.id})

def CostMonthlyView1(request, monthly_id):
    costs = MonthlyCosts.objects.filter(monthly_id = monthly_id)
    return render(request, 'viewscostsMonthly1.html', {'costs': costs,'monthly_id':monthly_id, 'number': Monthly.objects.get(pk=monthly_id).number})

def FinCostMonthlyCreate(request, monthly_id):
    month = Monthly.objects.get(pk=monthly_id)
    form = indexcost()
    if request.method == 'POST':
        actualize(monthly_id)
        form = indexcost(request.POST)
        if form.is_valid():
            name = form.cleaned_data['name']
            volume = form.cleaned_data['volume']
            currency = form.cleaned_data['currency']
            item = FinCosts(monthly=month,name=name,volume=volume, currency=currency)
            item.save()
            actualize(monthly_id)
        return redirect ('FinCostMonthlyView', monthly_id)
    context = {
        'form': form,
    }
    return render(request, 'FinformsCostMonthly.html', context)

def FinCostMonthlyDelete(request, cost_id):
    item = FinCosts.objects.get(id = cost_id)
    month = Monthly.objects.get(pk=item.monthly_id)
    all = FinCosts.objects.filter(monthly_id = item.monthly_id )
    print(all)
    if request.method == 'POST':
        item.delete()
        actualize(item.monthly_id)
        return redirect ('FinCostMonthlyView', month.id)
    return render(request, 'FinConfirmationMonthly.html',{'id': month.id})

def FinCostMonthlyView(request, monthly_id):
    costs = FinCosts.objects.filter(monthly_id = monthly_id)
    return render(request, 'FinviewscostsMonthly.html', {'costs': costs, 'monthly_id': monthly_id})

def FinCostMonthlyCreate1(request, monthly_id):
    month = Monthly.objects.get(pk=monthly_id)
    form = indexcost()
    if request.method == 'POST':
        actualize(monthly_id)
        form = indexcost(request.POST)
        if form.is_valid():
            name = form.cleaned_data['name']
            volume = form.cleaned_data['volume']
            currency = form.cleaned_data['currency']
            item = FinCosts(monthly=month,name=name,volume=volume, currency=currency)
            item.save()
            actualize(monthly_id)
        return redirect ('FinCostMonthlyView1', monthly_id)
    context = {
        'form': form,
        'id': monthly_id
    }
    return render(request, 'FinformsCostMonthly1.html', context)

def FinCostMonthlyDelete1(request, cost_id):
    item = FinCosts.objects.get(id = cost_id)
    month = Monthly.objects.get(pk=item.monthly_id)
    all = FinCosts.objects.filter(monthly_id = item.monthly_id )
    print(all)
    if request.method == 'POST':
        item.delete()
        actualize(item.monthly_id)
        return redirect ('FinCostMonthlyView1', month.id)
    return render(request, 'FinConfirmationMonthly1.html',{'id': month.id})

def FinCostMonthlyView1(request, monthly_id):
    costs = FinCosts.objects.filter(monthly_id = monthly_id)
    return render(request, 'FinviewscostsMonthly1.html', {'costs': costs, 'monthly_id': monthly_id, 'number': Monthly.objects.get(pk=monthly_id).number})

def CostDelete(request, cost_id):
    item = Costs.objects.get(id = cost_id)
    if request.method == 'POST':
        item.delete()
        actualizeShip(item.shipment_id)
        return redirect ('CostViews', item.shipment_id)
    return render(request, 'ConfirmationCost.html',{'id': item.shipment_id})

def CostViews(request, shipment_id):
    costs = Costs.objects.filter(shipment_id = shipment_id)
    item = Shipment.objects.get(pk = shipment_id)
    city = Ports.objects.filter(port=item.po.Origin)[0].country
    if request.user.username[:6] == 'import':
        city = Ports.objects.filter(port=item.po.so.destination)[0].country
    return render(request, 'viewscosts.html', {'costs': costs, 'shipment_id': shipment_id,'city':city})

def C(request, shipment_id):
    a = Shipment.objects.get(pk=shipment_id)
    costs = Costs.objects.filter(shipment_id = shipment_id)
    return render(request, 'costs.html', {'costs': costs, 'number': a.number, 'shipment_id':shipment_id})

def CostDelete1(request, cost_id):
    item = Costs.objects.get(id = cost_id)
    if request.method == 'POST':
        item.delete()
        actualizeShip(item.shipment_id)
        return redirect ('C', item.shipment_id)
    return render(request, 'ConfirmationCost1.html', {'id': item.shipment_id})

def Cost(request, shipment_id):
    a = Shipment.objects.get(pk=shipment_id)
    costs = Costs.objects.filter(shipment_id = shipment_id)
    return render(request, 'viewscosts1.html', {'costs': costs, 'shipment_id': a.po_id})

def ShowFreight(request, readiness_id):

    read = Readiness.objects.get(pk=readiness_id)
    country = Ports.objects.filter(port=read.Origin)[0].country
    POL = read.po.Origin
    POD = read.po.so.destination
    options = Freight.objects.filter(POL=POL).filter(POD=POD)
    salesorder = read.po.so

    for i in options:
        try:
            response = requests.get(f'https://www.cbr-xml-daily.ru/daily_json.js')
            rate = response.json()['Valute']['USD']['Value'] / response.json()['Valute']['EUR']['Value']
            rate = float('{:.6f}'.format(rate))

        except:
            rate = 0.88

        if read.po.so.currency == 'EUR':
            sale = float(read.po.so.cost) / float(rate)
        else:
            sale = float(read.po.so.cost)

        if read.po.currency == 'EUR':
            purchaise = float(read.po.price) / float(rate)
        else:
            purchaise = float(read.po.price)

        if i.currencyrate == 'EUR':
            rate = float(i.rate) / float(rate)
        else:
            rate = float(i.rate)

        if i.currencyadd == 'EUR':
            additional = float(i.additional) / float(rate)
        else:
            additional = float(i.additional)

        pcd = 0
        if salesorder.currency == 'USD':
            pdc = -4/36500 * (float(salesorder.cpt)*float(salesorder.cost)*float(salesorder.min))
        else:
            pdc = -4/36500 * (float(salesorder.cpt) * float(salesorder.cost)/float(rate) * float(salesorder.min))

        margin = sale - purchaise - rate / float(read.po.so.min) - float(additional) / float(read.cntr * read.po.so.min) - float(pdc) / float(read.po.so.min)
        i.margin = float('{:.2f}'.format(margin))
        i.save()


    return render(request, 'ShowFreight.html', {'price': options,'POL': POL,'POD': POD,'country':country,'readiness_id':readiness_id})

def actualizeShip(shipment_id):
    cost = Costs.objects.filter(shipment_id=shipment_id)
    ship = Shipment.objects.get(pk=shipment_id)

    rate = ShipmentRate.objects.get(shipment=ship)

    order = PO.objects.get(pk = ship.po_id)
    sale = SO.objects.get(pk=order.so_id)

    a = 0

    for i in cost:
        if i.name == 'Sale' and i.currency == 'USD':
            a += float(i.volume)
        elif i.name == 'Sale' and i.currency == 'EUR':
            a += float(i.volume)/float(rate.rate)

        elif i.name == 'Purchaise' and i.currency == 'USD':
            a += float(i.volume)

        elif i.name == 'Purchaise' and i.currency == 'EUR':
            a += float(i.volume)/float(rate.rate)

        elif i.currency == 'USD':
            a += float(i.volume) / float(sale.min)

        elif i.currency == 'EUR':
            a += ((float(i.volume) / float(sale.min)))/float(rate.rate)


    ship.margin = float('{:.2f}'.format(a))
    ship.save()
    result = a*float(rate.rate)
    print('USD',ship.margin)

    ship.marginEUR = float('{:.2f}'.format(result))
    ship.save()
    print(ship.marginEUR)

def Select(request, freight_id, readiness_id):
    readiness = Readiness.objects.get(pk=readiness_id)
    form = shipmentforma()
    if readiness.po == 'number':
        return redirect('OPS')

    if request.method == 'POST':

        freight = Freight.objects.get(pk=freight_id)
        readiness = Readiness.objects.get(pk=readiness_id)
        order = PO.objects.get(pk=readiness.po_id)
        salesorder = SO.objects.get(pk=order.so_id)

        obj = Shipment(po = order,number = '', forwarder = freight.forwarder,carrier = freight.Line,cntr=readiness.cntr,bknumber = '',ETD='',ETA='',margin=0,marginEUR=0,BK=False,SI=False,Magic=False, Truck = False,shipinstr='',equip='')

        form = shipmentforma(request.POST,instance = obj)
        if form.is_valid():
            item = form.save()
            item.ETD = item.ETD[8] + item.ETD[9] + '.' + item.ETD[5] + item.ETD[6] + '.' + item.ETD[0] + item.ETD[1] + item.ETD[2] + item.ETD[3]
            item.save()
            item.ETA = item.ETA[8] + item.ETA[9] + '.' + item.ETA[5] + item.ETA[6] + '.' + item.ETA[0] + item.ETA[1] + item.ETA[2] + item.ETA[3]
            item.save()
            try:
                response = requests.get(f'https://www.cbr-xml-daily.ru/daily_json.js')
                rate = response.json()['Valute']['USD']['Value'] / response.json()['Valute']['EUR']['Value']
                rate = float('{:.6f}'.format(rate))
            except:
                rate = 0.88

            var = ShipmentRate(shipment=item,rate=rate)
            var.save()
            readiness.delete()

            cost2 = Costs(shipment=item, name='Sale', volume=salesorder.cost, currency=salesorder.currency)
            cost2.save()
            cost3 = Costs(shipment=item, name='Purchaise', volume=-order.price, currency=order.currency)
            cost3.save()

            cost = Costs(shipment=item, name='Freight', volume=-float(freight.rate), currency=freight.currencyrate)
            cost.save()
            cost1 = Costs(shipment=item, name='Extra', volume=-float(freight.additional) / float(readiness.cntr), currency=freight.currencyadd)
            cost1.save()

            pcd = 0
            if salesorder.currency == 'USD':
                pdc = -4/36500 * (float(salesorder.cpt)*float(salesorder.cost)*float(salesorder.min))
            else:
                pdc = -4/36500 * (float(salesorder.cpt) * float(salesorder.cost)/float(rate) * float(salesorder.min))

            cost4 = Costs(shipment=item, name='Payment Dealay', volume=pdc, currency='USD')
            cost4.save()

            item.carrier = freight.Line
            item.save()

            actualizeShip(item.id)

            return redirect('OPS')

    context = {
        'form': form,
        'PO': readiness.po_id
    }
    return render(request, 'formsShipment.html', context)

def Skip(request, readiness_id):

    readiness = Readiness.objects.get(pk=readiness_id)

    order = PO.objects.get(pk=readiness.po_id)
    salesorder = SO.objects.get(pk=order.so_id)

    item = Shipment(po=order, number=readiness.number + '-xxx', forwarder='', carrier='', cntr=1,
                   bknumber='', ETD='', ETA='', margin=0, marginEUR=0, BK=False, SI=False, Magic=False,
                   Truck=True,equip='Truck',shipinstr='')
    item.save()
    try:
        response = requests.get(f'https://www.cbr-xml-daily.ru/daily_json.js')
        rate = response.json()['Valute']['USD']['Value'] / response.json()['Valute']['EUR']['Value']
        rate = float('{:.6f}'.format(rate))
    except:
        rate = 0.88

    var = ShipmentRate(shipment=item, rate=rate)
    var.save()
    readiness.delete()

    cost2 = Costs(shipment=item, name='Sale', volume=salesorder.cost, currency=salesorder.currency)
    cost2.save()
    cost3 = Costs(shipment=item, name='Purchaise', volume=-order.price, currency=order.currency)
    cost3.save()

    pcd = 0
    if salesorder.currency == 'USD':
        pdc = -4 / 36500 * (float(salesorder.cpt) * float(salesorder.cost) * float(salesorder.min))
    else:
        pdc = -4 / 36500 * (float(salesorder.cpt) * float(salesorder.cost) / float(rate) * float(salesorder.min))

    cost4 = Costs(shipment=item, name='Payment Dealay', volume=pdc, currency='USD')
    cost4.save()

    actualizeShip(item.id)

    now = datetime.now(timezone.utc)
    m = counterupd.objects.get(index='1')

    new = Containers(us = request.user,shipment=item,number = '',bales=0,gross = 0, tara=0,vgm = 0)
    new.save()

    nuevo = counter.objects.get(name=item.po.Origin.country)
    nuevo.volume += 1
    nuevo.save()
    print(nuevo.volume)

    nuevo1 = counter.objects.get(name=request.user.username)
    nuevo1.volume += 1
    nuevo1.save()

    return redirect('OPS')

def Spot(request, readiness_id):

    readiness = Readiness.objects.get(pk=readiness_id)

    order = PO.objects.get(pk=readiness.po_id)
    salesorder = SO.objects.get(pk=order.so_id)

    item = Shipment(po=order, number=readiness.number + '-xxx', forwarder='', carrier='', cntr=readiness.cntr,
                   bknumber='', ETD='', ETA='', margin=0, marginEUR=0, BK=False, SI=False, Magic=False,
                   Truck=False,equip='40 HQ',shipinstr='')
    item.save()
    try:
        response = requests.get(f'https://www.cbr-xml-daily.ru/daily_json.js')
        rate = response.json()['Valute']['USD']['Value'] / response.json()['Valute']['EUR']['Value']
        rate = float('{:.6f}'.format(rate))
    except:
        rate = 0.88

    var = ShipmentRate(shipment=item, rate=rate)
    var.save()
    readiness.delete()

    cost2 = Costs(shipment=item, name='Sale', volume=salesorder.cost, currency=salesorder.currency)
    cost2.save()
    cost3 = Costs(shipment=item, name='Purchaise', volume=-order.price, currency=order.currency)
    cost3.save()

    pcd = 0
    if salesorder.currency == 'USD':
        pdc = -4 / 36500 * (float(salesorder.cpt) * float(salesorder.cost) * float(salesorder.min))
    else:
        pdc = -4 / 36500 * (float(salesorder.cpt) * float(salesorder.cost) / float(rate) * float(salesorder.min))

    cost4 = Costs(shipment=item, name='Payment Dealay', volume=pdc, currency='USD')
    cost4.save()

    actualizeShip(item.id)

    return redirect('OPS')